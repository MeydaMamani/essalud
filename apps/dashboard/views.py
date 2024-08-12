
from django.http import JsonResponse, HttpResponse, QueryDict
from django.core import serializers

from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.views.generic import TemplateView, FormView, View

from django.contrib.auth import authenticate, login, logout, get_user_model
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.http import HttpResponseRedirect

from datetime import date, datetime
from django.db import connection
import json
import locale

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color

User = get_user_model()
from apps.person.models import Person
from .forms import LoginForm


# Create your views here.
class HomeView(TemplateView):
    template_name = 'base.html'


class LoginView(FormView):
    template_name = 'login.html'
    form_class = LoginForm
    success_url = reverse_lazy('dashboard:dash')

    @method_decorator(csrf_protect)
    @method_decorator(never_cache)

    #verifica la petición
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return HttpResponseRedirect(self.get_success_url())
        else:
            return super(LoginView, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        user = authenticate(
            username=form.cleaned_data['username'],
            password=form.cleaned_data['password']
        )
        login(self.request, user)

        try:
            ObjPerson = Person.objects.get(pk=user.id_person.id)
            ObjUser = User.objects.get(pk=user.pk)
            self.request.session['sytem'] = {'eid': ObjPerson.eid_id, 'full_name': ObjPerson.last_name0+' '+ObjPerson.last_name1+', '+ObjPerson.names.title(),
                                            'doc': ObjPerson.pdoc, 'red': ObjUser.id_red.pk, 'redCode': ObjUser.id_red.code, 'redLevel': ObjUser.id_red.level,
                                            'redName': ObjUser.id_red.name,
                                            'redState': ObjUser.id_red.state
                                        }

        except:
            print("Hay un error en los valores de entrada")

        return super(LoginView, self).form_valid(form)

    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     context['form3'] = ChangePassForm
    #     return context

def logoutUser(request):
    logout(request)
    return HttpResponseRedirect('/')


class DahboardView(TemplateView):
    template_name = 'dash.html'


class AttentionToday(View):
    def get(self, request, *args, **kwargs):
        mes = request.GET['mes']
        if len(request.GET['mes']) == 1:
            mes = '0'+request.GET['mes']
        else:
            mes = request.GET['mes']

        a = connection.cursor()
        a.execute("""SELECT establecimiento, ape_nombres, fec_nac, documento, Convert(Integer, Datediff(Day, [fec_nac], EOMONTH('%s-%s-01'))/30) Edad,
                    CASE WHEN CRED11 IS NOT NULL THEN CONCAT(CRED11, ' / CRED11')
                        WHEN CRED10 IS NOT NULL THEN CONCAT(CRED10, ' / CRED10')
                        WHEN CRED9 IS NOT NULL THEN CONCAT(CRED9, ' / CRED9')
                        WHEN CRED8 IS NOT NULL THEN CONCAT(CRED8, ' / CRED8')
                        WHEN CRED7 IS NOT NULL THEN CONCAT(CRED7, ' / CRED7')
                        WHEN CRED6 IS NOT NULL THEN CONCAT(CRED6, ' / CRED6')
                        WHEN CRED5 IS NOT NULL THEN CONCAT(CRED5, ' / CRED5')
                        WHEN CRED4 IS NOT NULL THEN CONCAT(CRED4, ' / CRED4')
                        WHEN CRED3 IS NOT NULL THEN CONCAT(CRED3, ' / CRED3')
                        WHEN CRED2 IS NOT NULL THEN CONCAT(CRED2, ' / CRED2')
                        WHEN CRED1 IS NOT NULL THEN CONCAT(CRED1, ' / CRED1')
                        WHEN ctrl4rn IS NOT NULL THEN CONCAT(ctrl4rn, ' / CRED4_RN')
                        WHEN ctrl3rn IS NOT NULL THEN CONCAT(ctrl3rn, ' / CRED3_RN')
                        WHEN ctrl2rn IS NOT NULL THEN CONCAT(ctrl2rn, ' / CRED2_RN')
                        WHEN ctrl1rn IS NOT NULL THEN CONCAT(ctrl1rn, ' / CRED1_RN')
                    END ULT_CRED,
                    CASE WHEN suple11 IS NOT NULL THEN CONCAT(suple11, ' / SUPLE11')
                        WHEN suple10 IS NOT NULL THEN CONCAT(suple10, ' / SUPLE10')
                        WHEN suple9 IS NOT NULL THEN CONCAT(suple9, ' / SUPLE9')
                        WHEN suple8 IS NOT NULL THEN CONCAT(suple8, ' / SUPLE8')
                        WHEN suple7 IS NOT NULL THEN CONCAT(suple7, ' / SUPLE7')
                        WHEN suple6 IS NOT NULL THEN CONCAT(suple6, ' / SUPLE6')
                        WHEN suple5 IS NOT NULL THEN CONCAT(suple5, ' / SUPLE5')
                        WHEN suple4 IS NOT NULL THEN CONCAT(suple4, ' / SUPLE4')
                    ELSE null end ULT_SUPLE,
                    neumo2, rota2, polio2, penta2, neumo4, rota4, penta4, polio4, polio6, penta6,
                    CASE WHEN CRED1 IS NULL THEN 'CRED1' END C1,
                        CASE WHEN CRED2 IS NULL THEN 'CRED2' END C2,
                        CASE WHEN CRED3 IS NULL THEN 'CRED3' END C3,
                        CASE WHEN CRED4 IS NULL THEN 'CRED4' END C4,
                        CASE WHEN CRED5 IS NULL THEN 'CRED5' END C5,
                        CASE WHEN CRED6 IS NULL THEN 'CRED6' END C6,
                        CASE WHEN CRED7 IS NULL THEN 'CRED7' END C7,
                        CASE WHEN CRED8 IS NULL THEN 'CRED8' END C8,
                        CASE WHEN CRED9 IS NULL THEN 'CRED9' END C9,
                        CASE WHEN CRED10 IS NULL THEN 'CRED10' END C10,
                        CASE WHEN CRED11 IS NULL THEN 'CRED11' END C11,
                        CASE WHEN SUPLE4 IS NULL THEN 'SPL4' END SPL4,
                        CASE WHEN SUPLE5 IS NULL THEN 'SPL5' END SPL5,
                        CASE WHEN SUPLE6 IS NULL THEN 'SPL6' END SPL6,
                        CASE WHEN SUPLE7 IS NULL THEN 'SPL7' END SPL7,
                        CASE WHEN SUPLE8 IS NULL THEN 'SPL8' END SPL8,
                        CASE WHEN SUPLE9 IS NULL THEN 'SPL9' END SPL9,
                        CASE WHEN SUPLE10 IS NULL THEN 'SPL10' END SPL10,
                        CASE WHEN SUPLE11 IS NULL THEN 'SPL11' END SPL11,
                        CASE WHEN neumo2 IS NULL THEN 'NEUMO2M' END neu2,
                        CASE WHEN rota2 IS NULL THEN 'ROTA2M' END rot2,
                        CASE WHEN polio2 IS NULL THEN 'POLIO2M' END pol2,
                        CASE WHEN penta2 IS NULL THEN 'PENTA2M' END pen2,
                        CASE WHEN neumo4 IS NULL THEN 'NEUMO4M' END neu4,
                        CASE WHEN rota4 IS NULL THEN 'ROTA4M' END rot4,
                        CASE WHEN penta4 IS NULL THEN 'PENTA4M' END pen4,
                        CASE WHEN polio4 IS NULL THEN 'POLIO4M' END pol4,
                        CASE WHEN polio6 IS NULL THEN 'POLIO6M' END pol6,
                        CASE WHEN penta6 IS NULL THEN 'PENTA6M' END pen6
                    into ESSALUD.dbo.ult_aten
                    FROM packages_packchildfollow""" % (request.GET['anio'], mes))

        a.execute("""SELECT establecimiento, ape_nombres, documento, fec_nac, Edad,
                    CASE WHEN ULT_CRED IS NULL THEN
                        CASE WHEN Edad=1 THEN CONCAT(DATEADD(day, 30, fec_nac), ' / CRED1')
                            WHEN Edad=2 THEN CONCAT(DATEADD(day, 60, fec_nac), ' / CRED2')
                            WHEN Edad=3 THEN CONCAT(DATEADD(day, 90, fec_nac), ' / CRED3')
                            WHEN Edad=4 THEN CONCAT(DATEADD(day, 120, fec_nac), ' / CRED4')
                            WHEN Edad=5 THEN CONCAT(DATEADD(day, 150, fec_nac), ' / CRED5')
                            WHEN Edad=6 THEN CONCAT(DATEADD(day, 180, fec_nac), ' / CRED6')
                            WHEN Edad=7 THEN CONCAT(DATEADD(day, 210, fec_nac), ' / CRED7')
                            WHEN Edad=8 THEN CONCAT(DATEADD(day, 240, fec_nac), ' / CRED8')
                            WHEN Edad=9 THEN CONCAT(DATEADD(day, 270, fec_nac), ' / CRED9')
                            WHEN Edad=10 THEN CONCAT(DATEADD(day, 300, fec_nac), ' / CRED10')
                            WHEN Edad>=11 THEN CONCAT(DATEADD(day, 330, fec_nac), ' / CRED11')
                        END
                    ELSE
                    CASE WHEN Edad=1 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 60, fec_nac), ' / CRED1') END
                    WHEN Edad=2 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)),' / CRED2' )
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 60, fec_nac), ' / CRED2') END
                    WHEN Edad=3 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED3')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED3')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 90, fec_nac), ' / CRED3') END
                    WHEN Edad=4 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED4')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED4')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED4')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 120, fec_nac), ' / CRED4') END
                    WHEN Edad=5 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 150, fec_nac), ' / CRED5') END
                    WHEN Edad=6 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 180, fec_nac), ' / CRED6') END
                    WHEN Edad=7 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 210, fec_nac), ' / CRED7') END
                    WHEN Edad=8 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 240, fec_nac), ' / CRED8') END
                    WHEN Edad=9 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 270, fec_nac), ' / CRED9') END
                    WHEN Edad=10 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED10' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 270, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 300, fec_nac), ' / CRED10') END
                    WHEN Edad=11 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED11' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED10' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 270, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 300, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 330, fec_nac), ' / CRED11') END
                    WHEN Edad>11 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED11' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED10' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 270, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 300, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 330, fec_nac), ' / CRED11') END
                    END
                END cred_hoy,
                CASE WHEN ULT_SUPLE IS NULL THEN
                    CASE WHEN Edad=4 THEN CONCAT(DATEADD(day, 120, fec_nac), ' / SUPLE4')
                        WHEN Edad=5 THEN CONCAT(DATEADD(day, 150, fec_nac), ' / SUPLE5')
                        WHEN Edad=6 THEN CONCAT(DATEADD(day, 180, fec_nac), ' / SUPLE6')
                        WHEN Edad=7 THEN CONCAT(DATEADD(day, 210, fec_nac), ' / SUPLE7')
                        WHEN Edad=8 THEN CONCAT(DATEADD(day, 240, fec_nac), ' / SUPLE8')
                        WHEN Edad=9 THEN CONCAT(DATEADD(day, 270, fec_nac), ' / SUPLE9')
                        WHEN Edad=10 THEN CONCAT(DATEADD(day, 300, fec_nac), ' / SUPLE10')
                        WHEN Edad>=11 THEN CONCAT(DATEADD(day, 330, fec_nac), ' / SUPLE11')
                        WHEN Edad in (0,1,2,3) THEN 'NO TOCA'
                    END
                ELSE
                    CASE WHEN Edad=4 THEN CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN ULT_SUPLE END
                    WHEN Edad=5 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE5') END
                    WHEN Edad=6 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE6')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE6') END
                    WHEN Edad=7 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE7')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE7')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE7') END
                    WHEN Edad=8 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8') END
                    WHEN Edad=9 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE9' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9') END
                    WHEN Edad=10 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE10' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE9' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10') END
                    WHEN Edad>=11 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE11' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE10' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE9' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11') END
                    END
                END suple_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN neumo2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / NEUMO2')
                            WHEN neumo2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN neumo4 is not null then 'TIENE'
                            WHEN neumo2 is not null then CONCAT(DATEADD(day, 60, neumo2), ' / NEUMO4')
                            WHEN neumo2 is null and neumo4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / NEUMO4') end
                    WHEN Edad not in (2,4) then 'NO TOCA'
                END neumo_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN rota2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / ROTA2')
                            WHEN rota2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN rota4 is not null then 'TIENE'
                            WHEN rota2 is not null then CONCAT(DATEADD(day, 60, rota2), ' / ROTA4')
                            WHEN rota2 is null and rota4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / ROTA4') end
                    WHEN Edad not in (2,4) then 'NO TOCA'
                END rota_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN penta2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / PENTA2')
                            WHEN penta2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN penta4 is not null then 'TIENE'
                            WHEN penta2 is not null then CONCAT(DATEADD(day, 60, penta2), ' / PENTA4')
                            WHEN penta2 is null and penta4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / PENTA4') end
                    WHEN Edad=6 THEN
                        CASE WHEN polio6 is not null then 'TIENE'
                            WHEN polio4 is not null then CONCAT(DATEADD(day, 60, polio4), ' / PENTA6')
                            WHEN polio2 is not null then CONCAT(DATEADD(day, 120, polio2), ' / PENTA6')
                            WHEN polio2 is null and polio4 is null and polio6 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / PENTA6') end
                    WHEN Edad not in (2,4,6) then 'NO TOCA'
                END penta_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN polio2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / POLIO2')
                            WHEN polio2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN polio4 is not null then 'TIENE'
                            WHEN polio2 is not null then CONCAT(DATEADD(day, 60, polio2), ' / POLIO4')
                            WHEN polio2 is null and polio4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / POLIO4') end
                    WHEN Edad=6 THEN
                        CASE WHEN polio6 is not null then 'TIENE'
                            WHEN polio4 is not null then CONCAT(DATEADD(day, 60, polio4), ' / POLIO6')
                            WHEN polio2 is not null then CONCAT(DATEADD(day, 120, polio2), ' / POLIO6')
                            WHEN polio2 is null and polio4 is null and polio6 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / POLIO6') end
                    WHEN Edad not in (2,4,6) then 'NO TOCA'
                END polio_hoy,
                CASE
                    WHEN Edad=1 THEN C1
                    WHEN Edad=2 THEN TRIM(CONCAT(C1, ' ', C2))
                    WHEN Edad=3 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3))
                    WHEN Edad=4 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4))
                    WHEN Edad=5 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5))
                    WHEN Edad=6 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6))
                    WHEN Edad=7 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7))
                    WHEN Edad=8 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8))
                    WHEN Edad=9 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8, ' ', C9))
                    WHEN Edad=10 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8, ' ', C9, ' ', C10))
                    WHEN Edad>=11 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8, ' ', C9, ' ', C10, ' ', C11))
                ELSE NULL END cred_falta,
                CASE
                    WHEN Edad=4 THEN TRIM(SPL4)
                    WHEN Edad=5 THEN TRIM(CONCAT(SPL4, ' ', SPL5))
                    WHEN Edad=6 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6))
                    WHEN Edad=7 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7))
                    WHEN Edad=8 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8))
                    WHEN Edad=9 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8, ' ', SPL9))
                    WHEN Edad=10 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8, ' ', SPL9, ' ', SPL10))
                    WHEN Edad>=11 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8, ' ', SPL9, ' ', SPL10, ' ', SPL11))
                    WHEN Edad in (0,1,2,3) THEN 'NO TOCA'
                ELSE NULL END suple_falta,
                CASE
                    WHEN Edad=2 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2))
                    WHEN Edad=3 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2))
                    WHEN Edad=4 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2, ' ', neu4, ' ', rot4, ' ', pen4, ' ', pol4))
                    WHEN Edad=5 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2, ' ', neu4, ' ', rot4, ' ', pen4, ' ', pol4))
                    WHEN Edad>=6 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2, ' ', neu4, ' ', rot4, ' ', pen4, ' ', pol4, ' ', pol6, ' ', pen6))
                ELSE NULL END vac_falta
            INTO ESSALUD.dbo.aten_hoy
            FROM ESSALUD.dbo.ult_aten""")

        a.execute("""select *
                    from ESSALUD.dbo.aten_hoy
                    where Edad<12 AND ((CASE WHEN (cred_hoy!='TIENE') then SUBSTRING(cred_hoy, 1,7) else cred_hoy end='%s-%s') OR
                    (CASE WHEN (suple_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(suple_hoy, 1,7) else suple_hoy end='%s-%s') OR
                    (CASE WHEN (neumo_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(neumo_hoy, 1,7) else neumo_hoy end='%s-%s') OR
                    (CASE WHEN (rota_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(rota_hoy, 1,7) else rota_hoy end='%s-%s') OR
                    (CASE WHEN (penta_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(penta_hoy, 1,7) else penta_hoy end='%s-%s') OR
                    (CASE WHEN (polio_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(polio_hoy, 1,7) else polio_hoy end='%s-%s'))
                    drop table ESSALUD.dbo.ult_aten
                    drop table ESSALUD.dbo.aten_hoy""" % (request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes))

        data = []
        for dat in a.fetchall():
            datos = {'eess': dat[0], 'ape_nombres': dat[1], 'num_doc': dat[2], 'fech_nac': dat[3].strftime("%d/%m/%y"), 'edad': dat[4], 'credhoy':dat[5],
                     'suplehoy':dat[6], 'neumohoy': dat[7], 'rotahoy': dat[8], 'pentahoy': dat[9], 'poliohoy': dat[10], 'credfalta': dat[11], 'suplefalta': dat[12],
                     'vacfalta': dat[13],
                    }

            data.append(datos)

        return HttpResponse(json.dumps(data), content_type='application/json')


class PrintAttentionToday(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_TIME, 'es_ES')
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:O2", "medium", "305496")
        set_border(self, ws, "A4:O4", "medium", "203764")
        set_border(self, ws, "A6:O6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws.merge_cells('B2:O2')
        ws.row_dimensions[2].height = 23
        ws.row_dimensions[4].height = 20
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 33
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 40

        ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='305496')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'ESSALUD: Seguimiento paquete niño -' + nameMonth.upper() +  ' ' + request.GET['anio']

        ws.merge_cells('A4:O4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='203764')
        ws['A4'] = 'CODIFICACION: Cred Rn: 99381.01 - Cred Mes: 99381 - Vacuna Antipolio: 90712 - Vacuna Pentavalente: 90722 - Dx Anemia: D500, D508, D509, D649, D539 - Suplementación: 99199.17, 99199.19 - Prematuros: P073, P071, P0711, P00712 - Dosaje Hemoglobina: 85018, 85018.01'

        ws.merge_cells('A6:O6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: EsSalud con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A8'] = '#'
        ws['A8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['B8'] = 'Centro Asistencial'
        ws['B8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B8'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['C8'] = 'Documento'
        ws['C8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['D8'] = 'Apellidos y Nombres'
        ws['D8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D8'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['E8'] = 'Fecha Nacido'
        ws['E8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E8'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['F8'] = 'Edad'
        ws['F8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F8'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['G8'] = 'Cred Hoy'
        ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['H8'] = 'Suple Hoy'
        ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['I8'] = 'Neumo Hoy'
        ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['J8'] = 'Rota Hoy'
        ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['K8'] = 'Penta Hoy'
        ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['L8'] = 'Polio Hoy'
        ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['M8'] = 'Cred Falta'
        ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['M8'].fill = PatternFill(start_color='e8dad7', end_color='e8dad7', fill_type='solid')

        ws['N8'] = 'Suple Falta'
        ws['N8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N8'].fill = PatternFill(start_color='e8dad7', end_color='e8dad7', fill_type='solid')

        ws['O8'] = 'Vacunas Falta'
        ws['O8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['O8'].fill = PatternFill(start_color='e8dad7', end_color='e8dad7', fill_type='solid')

        if len(request.GET['mes']) == 1:
            mes = '0'+request.GET['mes']
        else:
            mes = request.GET['mes']

        a = connection.cursor()
        a.execute("""SELECT establecimiento, ape_nombres, fec_nac, documento, Convert(Integer, Datediff(Day, [fec_nac], EOMONTH('%s-%s-01'))/30) Edad,
                    CASE WHEN CRED11 IS NOT NULL THEN CONCAT(CRED11, ' / CRED11')
                        WHEN CRED10 IS NOT NULL THEN CONCAT(CRED10, ' / CRED10')
                        WHEN CRED9 IS NOT NULL THEN CONCAT(CRED9, ' / CRED9')
                        WHEN CRED8 IS NOT NULL THEN CONCAT(CRED8, ' / CRED8')
                        WHEN CRED7 IS NOT NULL THEN CONCAT(CRED7, ' / CRED7')
                        WHEN CRED6 IS NOT NULL THEN CONCAT(CRED6, ' / CRED6')
                        WHEN CRED5 IS NOT NULL THEN CONCAT(CRED5, ' / CRED5')
                        WHEN CRED4 IS NOT NULL THEN CONCAT(CRED4, ' / CRED4')
                        WHEN CRED3 IS NOT NULL THEN CONCAT(CRED3, ' / CRED3')
                        WHEN CRED2 IS NOT NULL THEN CONCAT(CRED2, ' / CRED2')
                        WHEN CRED1 IS NOT NULL THEN CONCAT(CRED1, ' / CRED1')
                        WHEN ctrl4rn IS NOT NULL THEN CONCAT(ctrl4rn, ' / CRED4_RN')
                        WHEN ctrl3rn IS NOT NULL THEN CONCAT(ctrl3rn, ' / CRED3_RN')
                        WHEN ctrl2rn IS NOT NULL THEN CONCAT(ctrl2rn, ' / CRED2_RN')
                        WHEN ctrl1rn IS NOT NULL THEN CONCAT(ctrl1rn, ' / CRED1_RN')
                    END ULT_CRED,
                    CASE WHEN suple11 IS NOT NULL THEN CONCAT(suple11, ' / SUPLE11')
                        WHEN suple10 IS NOT NULL THEN CONCAT(suple10, ' / SUPLE10')
                        WHEN suple9 IS NOT NULL THEN CONCAT(suple9, ' / SUPLE9')
                        WHEN suple8 IS NOT NULL THEN CONCAT(suple8, ' / SUPLE8')
                        WHEN suple7 IS NOT NULL THEN CONCAT(suple7, ' / SUPLE7')
                        WHEN suple6 IS NOT NULL THEN CONCAT(suple6, ' / SUPLE6')
                        WHEN suple5 IS NOT NULL THEN CONCAT(suple5, ' / SUPLE5')
                        WHEN suple4 IS NOT NULL THEN CONCAT(suple4, ' / SUPLE4')
                    ELSE null end ULT_SUPLE,
                    neumo2, rota2, polio2, penta2, neumo4, rota4, penta4, polio4, polio6, penta6,
                    CASE WHEN CRED1 IS NULL THEN 'CRED1' END C1,
                        CASE WHEN CRED2 IS NULL THEN 'CRED2' END C2,
                        CASE WHEN CRED3 IS NULL THEN 'CRED3' END C3,
                        CASE WHEN CRED4 IS NULL THEN 'CRED4' END C4,
                        CASE WHEN CRED5 IS NULL THEN 'CRED5' END C5,
                        CASE WHEN CRED6 IS NULL THEN 'CRED6' END C6,
                        CASE WHEN CRED7 IS NULL THEN 'CRED7' END C7,
                        CASE WHEN CRED8 IS NULL THEN 'CRED8' END C8,
                        CASE WHEN CRED9 IS NULL THEN 'CRED9' END C9,
                        CASE WHEN CRED10 IS NULL THEN 'CRED10' END C10,
                        CASE WHEN CRED11 IS NULL THEN 'CRED11' END C11,
                        CASE WHEN SUPLE4 IS NULL THEN 'SPL4' END SPL4,
                        CASE WHEN SUPLE5 IS NULL THEN 'SPL5' END SPL5,
                        CASE WHEN SUPLE6 IS NULL THEN 'SPL6' END SPL6,
                        CASE WHEN SUPLE7 IS NULL THEN 'SPL7' END SPL7,
                        CASE WHEN SUPLE8 IS NULL THEN 'SPL8' END SPL8,
                        CASE WHEN SUPLE9 IS NULL THEN 'SPL9' END SPL9,
                        CASE WHEN SUPLE10 IS NULL THEN 'SPL10' END SPL10,
                        CASE WHEN SUPLE11 IS NULL THEN 'SPL11' END SPL11,
                        CASE WHEN neumo2 IS NULL THEN 'NEUMO2M' END neu2,
                        CASE WHEN rota2 IS NULL THEN 'ROTA2M' END rot2,
                        CASE WHEN polio2 IS NULL THEN 'POLIO2M' END pol2,
                        CASE WHEN penta2 IS NULL THEN 'PENTA2M' END pen2,
                        CASE WHEN neumo4 IS NULL THEN 'NEUMO4M' END neu4,
                        CASE WHEN rota4 IS NULL THEN 'ROTA4M' END rot4,
                        CASE WHEN penta4 IS NULL THEN 'PENTA4M' END pen4,
                        CASE WHEN polio4 IS NULL THEN 'POLIO4M' END pol4,
                        CASE WHEN polio6 IS NULL THEN 'POLIO6M' END pol6,
                        CASE WHEN penta6 IS NULL THEN 'PENTA6M' END pen6
                    into ESSALUD.dbo.ult_aten
                    FROM packages_packchildfollow""" % (request.GET['anio'], mes))

        a.execute("""SELECT establecimiento, ape_nombres, documento, fec_nac, Edad,
                    CASE WHEN ULT_CRED IS NULL THEN
                        CASE WHEN Edad=1 THEN CONCAT(DATEADD(day, 30, fec_nac), ' / CRED1')
                            WHEN Edad=2 THEN CONCAT(DATEADD(day, 60, fec_nac), ' / CRED2')
                            WHEN Edad=3 THEN CONCAT(DATEADD(day, 90, fec_nac), ' / CRED3')
                            WHEN Edad=4 THEN CONCAT(DATEADD(day, 120, fec_nac), ' / CRED4')
                            WHEN Edad=5 THEN CONCAT(DATEADD(day, 150, fec_nac), ' / CRED5')
                            WHEN Edad=6 THEN CONCAT(DATEADD(day, 180, fec_nac), ' / CRED6')
                            WHEN Edad=7 THEN CONCAT(DATEADD(day, 210, fec_nac), ' / CRED7')
                            WHEN Edad=8 THEN CONCAT(DATEADD(day, 240, fec_nac), ' / CRED8')
                            WHEN Edad=9 THEN CONCAT(DATEADD(day, 270, fec_nac), ' / CRED9')
                            WHEN Edad=10 THEN CONCAT(DATEADD(day, 300, fec_nac), ' / CRED10')
                            WHEN Edad>=11 THEN CONCAT(DATEADD(day, 330, fec_nac), ' / CRED11')
                        END
                    ELSE
                    CASE WHEN Edad=1 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 60, fec_nac), ' / CRED1') END
                    WHEN Edad=2 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)),' / CRED2' )
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 60, fec_nac), ' / CRED2') END
                    WHEN Edad=3 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED3')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED3')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 90, fec_nac), ' / CRED3') END
                    WHEN Edad=4 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED4')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED4')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED4')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 120, fec_nac), ' / CRED4') END
                    WHEN Edad=5 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED5')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 150, fec_nac), ' / CRED5') END
                    WHEN Edad=6 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED6')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 180, fec_nac), ' / CRED6') END
                    WHEN Edad=7 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED7')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 210, fec_nac), ' / CRED7') END
                    WHEN Edad=8 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED8')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 240, fec_nac), ' / CRED8') END
                    WHEN Edad=9 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED9')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 270, fec_nac), ' / CRED9') END
                    WHEN Edad=10 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED10' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 270, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED10')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 300, fec_nac), ' / CRED10') END
                    WHEN Edad=11 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED11' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED10' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 270, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 300, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 330, fec_nac), ' / CRED11') END
                    WHEN Edad>11 THEN
                        CASE WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED11' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED10' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED9' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED8' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED7' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED6' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED5' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED4' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED3' THEN CONCAT(DATEADD(day, 240, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED2' THEN CONCAT(DATEADD(day, 270, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)='CRED1' THEN CONCAT(DATEADD(day, 300, CAST(SUBSTRING(ULT_CRED, 1, 10) AS date)), ' / CRED11')
                            WHEN SUBSTRING(ULT_CRED, 14, 25)!='CRED1' THEN CONCAT(DATEADD(day, 330, fec_nac), ' / CRED11') END
                    END
                END cred_hoy,
                CASE WHEN ULT_SUPLE IS NULL THEN
                    CASE WHEN Edad=4 THEN CONCAT(DATEADD(day, 120, fec_nac), ' / SUPLE4')
                        WHEN Edad=5 THEN CONCAT(DATEADD(day, 150, fec_nac), ' / SUPLE5')
                        WHEN Edad=6 THEN CONCAT(DATEADD(day, 180, fec_nac), ' / SUPLE6')
                        WHEN Edad=7 THEN CONCAT(DATEADD(day, 210, fec_nac), ' / SUPLE7')
                        WHEN Edad=8 THEN CONCAT(DATEADD(day, 240, fec_nac), ' / SUPLE8')
                        WHEN Edad=9 THEN CONCAT(DATEADD(day, 270, fec_nac), ' / SUPLE9')
                        WHEN Edad=10 THEN CONCAT(DATEADD(day, 300, fec_nac), ' / SUPLE10')
                        WHEN Edad>=11 THEN CONCAT(DATEADD(day, 330, fec_nac), ' / SUPLE11')
                        WHEN Edad in (0,1,2,3) THEN 'NO TOCA'
                    END
                ELSE
                    CASE WHEN Edad=4 THEN CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN ULT_SUPLE END
                    WHEN Edad=5 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE5') END
                    WHEN Edad=6 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE6')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE6') END
                    WHEN Edad=7 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE7')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE7')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE7') END
                    WHEN Edad=8 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE8') END
                    WHEN Edad=9 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE9' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE9') END
                    WHEN Edad=10 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE10' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE9' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE10') END
                    WHEN Edad>=11 THEN
                        CASE WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE11' THEN 'TIENE'
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE10' THEN CONCAT(DATEADD(day, 30, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE9' THEN CONCAT(DATEADD(day, 60, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE8' THEN CONCAT(DATEADD(day, 90, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE7' THEN CONCAT(DATEADD(day, 120, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE6' THEN CONCAT(DATEADD(day, 150, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE5' THEN CONCAT(DATEADD(day, 180, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11')
                            WHEN SUBSTRING(ULT_SUPLE, 14, 25)='SUPLE4' THEN CONCAT(DATEADD(day, 210, CAST(SUBSTRING(ULT_SUPLE, 1, 10) AS date)), ' / SUPLE11') END
                    END
                END suple_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN neumo2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / NEUMO2')
                            WHEN neumo2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN neumo4 is not null then 'TIENE'
                            WHEN neumo2 is not null then CONCAT(DATEADD(day, 60, neumo2), ' / NEUMO4')
                            WHEN neumo2 is null and neumo4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / NEUMO4') end
                    WHEN Edad not in (2,4) then 'NO TOCA'
                END neumo_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN rota2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / ROTA2')
                            WHEN rota2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN rota4 is not null then 'TIENE'
                            WHEN rota2 is not null then CONCAT(DATEADD(day, 60, rota2), ' / ROTA4')
                            WHEN rota2 is null and rota4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / ROTA4') end
                    WHEN Edad not in (2,4) then 'NO TOCA'
                END rota_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN penta2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / PENTA2')
                            WHEN penta2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN penta4 is not null then 'TIENE'
                            WHEN penta2 is not null then CONCAT(DATEADD(day, 60, penta2), ' / PENTA4')
                            WHEN penta2 is null and penta4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / PENTA4') end
                    WHEN Edad=6 THEN
                        CASE WHEN polio6 is not null then 'TIENE'
                            WHEN polio4 is not null then CONCAT(DATEADD(day, 60, polio4), ' / PENTA6')
                            WHEN polio2 is not null then CONCAT(DATEADD(day, 120, polio2), ' / PENTA6')
                            WHEN polio2 is null and polio4 is null and polio6 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / PENTA6') end
                    WHEN Edad not in (2,4,6) then 'NO TOCA'
                END penta_hoy,
                CASE WHEN Edad=2 THEN
                        CASE WHEN polio2 is null then CONCAT(DATEADD(day, 60, fec_nac), ' / POLIO2')
                            WHEN polio2 is not null then 'TIENE' end
                    WHEN Edad=4 THEN
                        CASE WHEN polio4 is not null then 'TIENE'
                            WHEN polio2 is not null then CONCAT(DATEADD(day, 60, polio2), ' / POLIO4')
                            WHEN polio2 is null and polio4 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / POLIO4') end
                    WHEN Edad=6 THEN
                        CASE WHEN polio6 is not null then 'TIENE'
                            WHEN polio4 is not null then CONCAT(DATEADD(day, 60, polio4), ' / POLIO6')
                            WHEN polio2 is not null then CONCAT(DATEADD(day, 120, polio2), ' / POLIO6')
                            WHEN polio2 is null and polio4 is null and polio6 is null then CONCAT(DATEADD(day, 120, fec_nac), ' / POLIO6') end
                    WHEN Edad not in (2,4,6) then 'NO TOCA'
                END polio_hoy,
                CASE
                    WHEN Edad=1 THEN C1
                    WHEN Edad=2 THEN TRIM(CONCAT(C1, ' ', C2))
                    WHEN Edad=3 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3))
                    WHEN Edad=4 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4))
                    WHEN Edad=5 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5))
                    WHEN Edad=6 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6))
                    WHEN Edad=7 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7))
                    WHEN Edad=8 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8))
                    WHEN Edad=9 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8, ' ', C9))
                    WHEN Edad=10 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8, ' ', C9, ' ', C10))
                    WHEN Edad>=11 THEN TRIM(CONCAT(C1, ' ', C2, ' ', C3, ' ', C4, ' ', C5, ' ', C6, ' ', C7, ' ', C8, ' ', C9, ' ', C10, ' ', C11))
                ELSE NULL END cred_falta,
                CASE
                    WHEN Edad=4 THEN TRIM(SPL4)
                    WHEN Edad=5 THEN TRIM(CONCAT(SPL4, ' ', SPL5))
                    WHEN Edad=6 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6))
                    WHEN Edad=7 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7))
                    WHEN Edad=8 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8))
                    WHEN Edad=9 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8, ' ', SPL9))
                    WHEN Edad=10 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8, ' ', SPL9, ' ', SPL10))
                    WHEN Edad>=11 THEN TRIM(CONCAT(SPL4, ' ', SPL5, ' ', SPL6, ' ', SPL7, ' ', SPL8, ' ', SPL9, ' ', SPL10, ' ', SPL11))
                    WHEN Edad in (0,1,2,3) THEN 'NO TOCA'
                ELSE NULL END suple_falta,
                CASE
                    WHEN Edad=2 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2))
                    WHEN Edad=3 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2))
                    WHEN Edad=4 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2, ' ', neu4, ' ', rot4, ' ', pen4, ' ', pol4))
                    WHEN Edad=5 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2, ' ', neu4, ' ', rot4, ' ', pen4, ' ', pol4))
                    WHEN Edad>=6 THEN TRIM(CONCAT(neu2, ' ', rot2, ' ', pol2, ' ', pen2, ' ', neu4, ' ', rot4, ' ', pen4, ' ', pol4, ' ', pol6, ' ', pen6))
                ELSE NULL END vac_falta
            INTO ESSALUD.dbo.aten_hoy
            FROM ESSALUD.dbo.ult_aten""")

        a.execute("""select *
                    from ESSALUD.dbo.aten_hoy
                    where Edad<12 AND ((CASE WHEN (cred_hoy!='TIENE') then SUBSTRING(cred_hoy, 1,7) else cred_hoy end='%s-%s') OR
                    (CASE WHEN (suple_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(suple_hoy, 1,7) else suple_hoy end='%s-%s') OR
                    (CASE WHEN (neumo_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(neumo_hoy, 1,7) else neumo_hoy end='%s-%s') OR
                    (CASE WHEN (rota_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(rota_hoy, 1,7) else rota_hoy end='%s-%s') OR
                    (CASE WHEN (penta_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(penta_hoy, 1,7) else penta_hoy end='%s-%s') OR
                    (CASE WHEN (polio_hoy not in ('TIENE', 'NO TOCA')) then SUBSTRING(polio_hoy, 1,7) else polio_hoy end='%s-%s'))
                    drop table ESSALUD.dbo.ult_aten
                    drop table ESSALUD.dbo.aten_hoy""" % (request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes, request.GET['anio'], mes))

        cont = 9
        num=1

        for pqt in a.fetchall():
            ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=1).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=1).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=1).value = num

            ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=cont, column=2).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=2).value = pqt[0]

            ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=3).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=3).value = pqt[2]

            ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=cont, column=4).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=4).value = pqt[1]

            ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=5).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=5).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=5).value = pqt[3]

            ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=6).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=6).value = pqt[4]

            ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=7).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=7).value = pqt[5]

            ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=8).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=8).value = pqt[6]

            ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=9).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=9).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=9).value = pqt[7]

            ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=10).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=10).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=10).value = pqt[8]

            ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=11).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=11).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=11).value = pqt[9]

            ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=cont, column=12).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=12).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=12).value = pqt[10]

            ws.cell(row=cont, column=13).alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
            ws.cell(row=cont, column=13).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=13).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=13).value = pqt[11]

            ws.cell(row=cont, column=14).alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
            ws.cell(row=cont, column=14).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=14).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=14).value = pqt[12]

            ws.cell(row=cont, column=15).alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
            ws.cell(row=cont, column=15).border = Border(bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=15).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=15).value = pqt[13]

            cont = cont+1
            num = num+1

        # sheet2 = wb.create_sheet('RESUMEN')
        # sheet2['A1'] = 'SUSCRIPCION'
        nombre_archivo = "DEIT_PASCO PAQUETE COMPLETO QUE LES TOCA A LA FECHA.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL PAQUETE NIÑO'
        wb.save(response)
        return response
