
# Create your views here.
from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView, View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, QueryDict
from datetime import date, datetime
import json
import locale

from django.db.models import Sum, F, FloatField, Case, When
from django.db.models.functions import Cast
from django.db import connection

from apps.packages.models import PackChildFollow, PregnantFollow
from apps.main.models import Provincia, Distrito, Establecimiento
from apps.follow_up.models import Anemia, VaccinexPat, Inmunization, Inmunization_C, InmunizationMinsa

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color


class KidsView(TemplateView):
    template_name = 'person/index.html'


class SearchKidsView(View):
    def get(self, request, *args, **kwargs):
        if request.GET['type'] == '1':
            data_saved = PackChildFollow.objects.filter(documento=request.GET['doc'])
            format_data = serializers.serialize('json', data_saved, indent=2, use_natural_foreign_keys=True)
            return HttpResponse(format_data, content_type='application/json')

        elif request.GET['type'] == '2':
            data_saved = PregnantFollow.objects.filter(documento=request.GET['doc'])
            format_data = serializers.serialize('json', data_saved, indent=2, use_natural_foreign_keys=True)
            return HttpResponse(format_data, content_type='application/json')

        elif request.GET['type'] == '3':
            data_saved = VaccinexPat.objects.filter(documento=request.GET['doc']).order_by('-fec_atencion')
            format_data = serializers.serialize('json', data_saved, indent=2, use_natural_foreign_keys=True)
            return HttpResponse(format_data, content_type='application/json')

        elif request.GET['type'] == '4':
            data_saved = Anemia.objects.filter(documento=request.GET['doc'])
            format_data = serializers.serialize('json', data_saved, indent=2, use_natural_foreign_keys=True)
            return HttpResponse(format_data, content_type='application/json')


class AnemiaKidsView(TemplateView):
    template_name = 'anemia/kids.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.session['sytem']['typeca'] == 'CA':
            context['establecimiento'] = Establecimiento.objects.filter(codigo=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DS':
            context['establecimiento'] = Establecimiento.objects.filter(dist_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'PR':
            context['establecimiento'] = Establecimiento.objects.filter(prov_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DP':
            context['establecimiento'] = Establecimiento.objects.filter(dep_id=self.request.session['sytem']['codeca'])
        return context


class NominalAnemia(View):
    def post(self, request, *args, **kwargs):
        dataList = []
        result = Anemia.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('establecimiento').annotate(menor=Sum(Case(When(num=1, then=1), default=0)),
                oneyear=Sum(Case(When(num=1, then=1), default=0)), twoyear=Sum(Case(When(num=1, then=2), default=0)))

        if request.POST['tipo'] == 'TODOS':
            if request.POST['eess'] == 'TODOS':
                if self.request.session['sytem']['typeca'] == 'CA':
                    totProv = Anemia.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

                elif self.request.session['sytem']['typeca'] == 'DS':
                    totProv = Anemia.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

                elif self.request.session['sytem']['typeca'] == 'PR':
                    totProv = Anemia.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

                elif self.request.session['sytem']['typeca'] == 'DP':
                    totProv = Anemia.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

            else:
                totProv = Anemia.objects.filter(cod_eess=request.POST['eess'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                totNominal = Anemia.objects.filter(cod_eess=request.POST['eess'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['eess'] == 'TODOS':
                if self.request.session['sytem']['typeca'] == 'CA':
                    totProv = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_eess=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_eess=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

                elif self.request.session['sytem']['typeca'] == 'DS':
                    totProv = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_dist=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_dist=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

                elif self.request.session['sytem']['typeca'] == 'PR':
                    totProv = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_prov=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_prov=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

                elif self.request.session['sytem']['typeca'] == 'DP':
                    totProv = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_dep=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                    totNominal = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_dep=self.request.session['sytem']['codeca'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                    totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

            else:
                totProv = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_eess=request.POST['eess'], anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(total=Sum('den'))
                totNominal = Anemia.objects.filter(grupo_edad=request.POST['tipo'], cod_eess=request.POST['eess'], anio=request.POST['anio'], mes=request.POST['mes']).order_by('cod_eess')
                totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

        dataList.append(list(totProv))
        dataList.append(totNominal)
        dataList.append(list(result))

        return HttpResponse(json.dumps(dataList), content_type='application/json')


class PrintNomAnem(View):
    def get(self, request, *args, **kwargs):
        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:Q2", "medium", "57267C")
        set_border(self, ws, "A4:Q4", "medium", "366092")
        set_border(self, ws, "A6:Q6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws.merge_cells('B2:X2')
        ws.row_dimensions[2].height = 23

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 33
        ws.column_dimensions['E'].width = 11
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['X'].width = 15

        ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO DE NIÑOS Y NIÑAS CON DX ANEMIA - ' + str(nameMonth).upper() + ' ' + str(request.GET['anio'])

        ws.merge_cells('A4:X4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['A4'] = 'CODIFICACION: '

        ws.merge_cells('A6:X6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A8'] = '#'
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['A8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['B8'] = 'Centro Asistencial'
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['B8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['C8'] = 'Documento'
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['C8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['D8'] = 'Apellidos y Nombres'
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['D8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['E8'] = 'Fecha Nacido'
        ws['E8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['E8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['F8'] = 'Dosaje 1'
        ws['F8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F8'].fill = PatternFill(start_color='C7ECF0', end_color='C7ECF0', fill_type='solid')
        ws['F8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['G8'] = 'Resultado 1'
        ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G8'].fill = PatternFill(start_color='C7ECF0', end_color='C7ECF0', fill_type='solid')
        ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['H8'] = 'Dosaje 2'
        ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H8'].fill = PatternFill(start_color='F0E6C7', end_color='F0E6C7', fill_type='solid')
        ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['I8'] = 'Resultado 2'
        ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I8'].fill = PatternFill(start_color='F0E6C7', end_color='F0E6C7', fill_type='solid')
        ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['J8'] = 'Dx Anemia 1'
        ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J8'].fill = PatternFill(start_color='F0C7E6', end_color='F0C7E6', fill_type='solid')
        ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['K8'] = 'Dx Anemia 2'
        ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K8'].fill = PatternFill(start_color='F0C7E6', end_color='F0C7E6', fill_type='solid')
        ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['L8'] = 'Nutrición 1'
        ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['M8'] = 'Nutrición 2'
        ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['N8'] = 'Nutrición 3'
        ws['N8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['N8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['O8'] = 'Nutrición 4'
        ws['O8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['O8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['P8'] = 'Nutrición 5'
        ws['P8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['P8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['Q8'] = 'Nutrición 6'
        ws['Q8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['Q8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['R8'] = 'Enf 1'
        ws['R8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R8'].fill = PatternFill(start_color='a4c0de', end_color='a4c0de', fill_type='solid')
        ws['R8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['S8'] = 'Enf 2'
        ws['S8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S8'].fill = PatternFill(start_color='a4c0de', end_color='a4c0de', fill_type='solid')
        ws['S8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['T8'] = 'Enf 3'
        ws['T8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T8'].fill = PatternFill(start_color='a4c0de', end_color='a4c0de', fill_type='solid')
        ws['T8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['U8'] = 'Enf 4'
        ws['U8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U8'].fill = PatternFill(start_color='a4c0de', end_color='a4c0de', fill_type='solid')
        ws['U8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['V8'] = 'Enf 5'
        ws['V8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V8'].fill = PatternFill(start_color='a4c0de', end_color='a4c0de', fill_type='solid')
        ws['V8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['W8'] = 'Enf 6'
        ws['W8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W8'].fill = PatternFill(start_color='a4c0de', end_color='a4c0de', fill_type='solid')
        ws['W8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['X8'] = 'Grupo'
        ws['X8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['X8'].fill = PatternFill(start_color='C7DBF0', end_color='C7DBF0', fill_type='solid')
        ws['X8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        if request.GET['tipo'] == 'TODOS':
            if request.GET['eess'] == 'TODOS':
                if self.request.session['sytem']['typeca'] == 'CA':
                    totNominal = Anemia.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DS':
                    totNominal = Anemia.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'PR':
                    totNominal = Anemia.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DP':
                    totNominal = Anemia.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')

            elif request.GET['eess'] != 'TODOS':
                totNominal = Anemia.objects.filter(cod_eess=request.GET['eess'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')

        else:
            if request.GET['eess'] == 'TODOS':
                if self.request.session['sytem']['typeca'] == 'CA':
                    totNominal = Anemia.objects.filter(grupo_edad=request.GET['tipo'], cod_eess=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DS':
                    totNominal = Anemia.objects.filter(grupo_edad=request.GET['tipo'], cod_dist=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'PR':
                    totNominal = Anemia.objects.filter(grupo_edad=request.GET['tipo'], cod_prov=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DP':
                    totNominal = Anemia.objects.filter(grupo_edad=request.GET['tipo'], cod_dep=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')

            elif request.GET['eess'] != 'TODOS':
                totNominal = Anemia.objects.filter(grupo_edad=request.GET['tipo'], cod_eess=request.GET['eess'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')

        totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

        cont = 9
        cant = len(totNominal)
        num = 1
        if cant > 0:
            for nom in totNominal:
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=1).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=1).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=1).value = num

                ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=2).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=2).value = nom['fields']['establecimiento']

                ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=3).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=3).value = nom['fields']['documento']

                ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=4).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=4).value = nom['fields']['ape_nombres']

                ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=5).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=6).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=6).value = nom['fields']['dosaje1']

                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=7).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=7).value = nom['fields']['result1']

                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=8).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=8).value = nom['fields']['dosaje2']

                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=9).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=9).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=9).value = nom['fields']['result2']

                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=10).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=10).value = nom['fields']['dx_anemia1']

                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=11).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=11).value = nom['fields']['dx_anemia2']

                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                if 0 if nom['fields']['nutricion6'] == None else nom['fields']['nutricion6'] < nom['fields']['dx_anemia1']:
                    ws.cell(row=cont, column=12).font = Font(name='Calibri', size=9, color='FE0220')
                else:
                    ws.cell(row=cont, column=12).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=12).value = nom['fields']['nutricion6']

                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                if 0 if nom['fields']['nutricion7'] == None else nom['fields']['nutricion7'] < nom['fields']['dx_anemia1']:
                    ws.cell(row=cont, column=13).font = Font(name='Calibri', size=9, color='FE0220')
                else:
                    ws.cell(row=cont, column=13).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=13).value = nom['fields']['nutricion7']

                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                if 0 if nom['fields']['nutricion8'] == None else nom['fields']['nutricion8'] < nom['fields']['dx_anemia1']:
                    ws.cell(row=cont, column=14).font = Font(name='Calibri', size=9, color='FE0220')
                else:
                    ws.cell(row=cont, column=14).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=14).value = nom['fields']['nutricion8']

                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                if 0 if nom['fields']['nutricion9'] == None else nom['fields']['nutricion9'] < nom['fields']['dx_anemia1']:
                    ws.cell(row=cont, column=15).font = Font(name='Calibri', size=9, color='FE0220')
                else:
                    ws.cell(row=cont, column=15).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=15).value = nom['fields']['nutricion9']

                ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=16).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                if 0 if nom['fields']['nutricion10'] == None else nom['fields']['nutricion10'] < nom['fields']['dx_anemia1']:
                    ws.cell(row=cont, column=16).font = Font(name='Calibri', size=9, color='FE0220')
                else:
                    ws.cell(row=cont, column=16).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=16).value = nom['fields']['nutricion10']

                ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=17).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                if 0 if nom['fields']['nutricion11'] == None else nom['fields']['nutricion11'] < nom['fields']['dx_anemia1']:
                    ws.cell(row=cont, column=17).font = Font(name='Calibri', size=9, color='FE0220')
                else:
                    ws.cell(row=cont, column=17).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=17).value = nom['fields']['nutricion11']

                ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=18).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=18).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=18).value = nom['fields']['enf6']

                ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=19).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=19).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=19).value = nom['fields']['enf7']

                ws.cell(row=cont, column=20).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=20).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=20).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=20).value = nom['fields']['enf8']

                ws.cell(row=cont, column=21).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=21).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=21).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=21).value = nom['fields']['enf9']

                ws.cell(row=cont, column=22).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=22).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=22).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=22).value = nom['fields']['enf10']

                ws.cell(row=cont, column=23).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=23).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=23).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=23).value = nom['fields']['enf11']

                ws.cell(row=cont, column=24).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=24).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=24).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=24).value = nom['fields']['grupo_edad']

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO SEGUIMIENTO DE NIÑOS CON DX ANEMIA.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL NIÑOS RN'
        wb.save(response)
        return response


class MetasPriorView(TemplateView):
    template_name = 'goals/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['establecimiento'] = Establecimiento.objects.all()
        return context


class ListMetasPrior(View):
    def get(self, request, *args, **kwargs):
        dataList = []
        nameEess = Establecimiento.objects.get(codigo=request.GET['eess'])

        a = connection.cursor()
        a.execute("""select (select sum(meta) from metas_priorizadas where cod_centro=%s and anio=%s) meta, sum(IIF(ene is null, 0, ene)) ene,
                    sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar, sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may,
                    sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul, sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set],
                    sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov, sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+
                    sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))
                    +sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0,
                    [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                    from ESSALUD.dbo.avance where cod_centro=%s and anio=%s""" % (request.GET['eess'], request.GET['anio'], request.GET['eess'], request.GET['anio']))

        resulTotal = { 'name': nameEess.nombre + ' - ' + request.GET['anio'] }

        for tot in a.fetchall():
            if tot[0] == 0 or tot[0] == None or tot[13] == None:
                avGeneralAnio = 0
                avEne = 0
                avFeb = 0
                avMar = 0
                avAbr = 0
                avMay = 0
                avJun = 0
                avJul = 0
                avAgo = 0
                avSet = 0
                avOct = 0
                avNov = 0
                avDic = 0
            else:
                avGeneralAnio = round((tot[13]/tot[0])*100, 1)
                avEne = round((tot[1]/(tot[0]/12))*100, 1)
                avFeb = round((tot[1]+tot[2])/((tot[0]/12)*2)*100, 1)
                avMar = round((tot[1]+tot[2]+tot[3])/((tot[0]/12)*3)*100, 1)
                avAbr = round((tot[1]+tot[2]+tot[3]+tot[4])/((tot[0]/12)*4)*100, 1)
                avMay = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5])/((tot[0]/12)*5)*100, 1)
                avJun = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6])/((tot[0]/12)*6)*100, 1)
                avJul = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7])/((tot[0]/12)*7)*100, 1)
                avAgo = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8])/((tot[0]/12)*8)*100, 1)
                avSet = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9])/((tot[0]/12)*9)*100, 1)
                avOct = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9]+tot[10])/((tot[0]/12)*10)*100, 1)
                avNov = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9]+tot[10]+tot[11])/((tot[0]/12)*11)*100, 1)
                avDic = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9]+tot[10]+tot[11]+tot[12])/((tot[0]/12)*12)*100, 1)

            total = { 'meta': tot[0], 'ene': tot[1], 'feb': tot[2], 'mar': tot[3], 'abr': tot[4],
                    'may': tot[5], 'jun': tot[6], 'jul': tot[7], 'ago': tot[8], 'set': tot[9],
                    'oct': tot[10], 'nov': tot[11], 'dic': tot[12], 'avance': tot[13], 'avporcent': avGeneralAnio }

            totalAv = { 'av_ene': avEne, 'av_feb': avFeb, 'av_mar': avMar, 'av_abr': avAbr,
                     'av_may': avMay, 'av_jun': avJun, 'av_jul': avJul, 'av_ago': avAgo, 'av_set': avSet,
                     'av_oct': avOct, 'av_nov': avNov, 'av_dic': avDic }

        b = connection.cursor()
        b.execute("""SELECT cod_act, sum(meta) meta into ESSALUD.dbo.metas_prior from metas_priorizadas where cod_centro=%s and anio=%s group by cod_act""" % (request.GET['eess'], request.GET['anio']))
        b.execute("""SELECT b.nombre, c.meta, sum(IIF(ene is null, 0, ene)) ene, sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar,
                    sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may, sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul,
                    sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set], sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov,
                    sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+
                    sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))+sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+
                    sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0, [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                    FROM ESSALUD.dbo.avance a left join ESSALUD.dbo.actividades b on a.cod_Act=b.codigo left join ESSALUD.dbo.metas_prior c on a.cod_act=c.cod_act
                    where a.cod_centro=%s and a.anio=%s group by b.nombre, c.meta
                    drop table ESSALUD.dbo.metas_prior""" % (request.GET['eess'], request.GET['anio']))

        dataAct = []
        for avAct in b.fetchall():
            totalAct = { 'nombre': avAct[0], 'meta': avAct[1], 'ene': avAct[2], 'feb': avAct[3], 'mar': avAct[4], 'abr': avAct[5],
                     'may': avAct[6], 'jun': avAct[7], 'jul': avAct[8], 'ago': avAct[9], 'set': avAct[10], 'oct': avAct[11],
                     'nov': avAct[12], 'dic': avAct[13], 'avance': avAct[14], 'avporcent': round((avAct[14]/avAct[1])*100, 1) }

            dataAct.append(totalAct)

        c = connection.cursor()
        c.execute("""SELECT a.cod_act, b.nombre from metas_priorizadas a left join actividades b on a.cod_act=b.codigo where cod_centro=%s and anio=%s
                        group by a.cod_act, b.nombre""" % (request.GET['eess'], request.GET['anio']))

        d = connection.cursor()
        e = connection.cursor()

        dataActDetail = []
        dataAvActDetail = []
        dataSubActiv = []
        for act in c.fetchall():
            e.execute("""select (select sum(meta) from metas_priorizadas where cod_centro=%s and anio=%s and cod_act=%s) meta, sum(IIF(ene is null, 0, ene)) ene,
                        sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar, sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may,
                        sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul, sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set],
                        sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov, sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+
                        sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))
                        +sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0,
                        [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                        from avance where cod_centro=%s and anio=%s and cod_act=%s""" % (request.GET['eess'], request.GET['anio'], "'"+act[0]+"'", request.GET['eess'], request.GET['anio'], "'"+act[0]+"'"))

            for totsub in e.fetchall():
                if totsub[0] == 0 or totsub[0] == None or totsub[13] == None:
                    avSubAct = 0
                    subActene = 0
                    subActfeb = 0
                    subActmar = 0
                    subActabr = 0
                    subActmay = 0
                    subActjun = 0
                    subActjul = 0
                    subActago = 0
                    subActset = 0
                    subActoct = 0
                    subActnov = 0
                    subActdic = 0
                else:
                    avSubAct = round((totsub[13]/totsub[0])*100, 1)
                    subActene = round((0 if totsub[1] == None else totsub[1]/(totsub[0]/12))*100, 1)
                    subActfeb = round((0 if totsub[1] == None else totsub[1]+totsub[2])/((totsub[0]/12)*2)*100, 1)
                    subActmar = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3])/((totsub[0]/12)*3)*100, 1)
                    subActabr = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4])/((totsub[0]/12)*4)*100, 1)
                    subActmay = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5])/((totsub[0]/12)*5)*100, 1)
                    subActjun = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6])/((totsub[0]/12)*6)*100, 1)
                    subActjul = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7])/((totsub[0]/12)*7)*100, 1)
                    subActago = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8])/((totsub[0]/12)*8)*100, 1)
                    subActset = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9])/((totsub[0]/12)*9)*100, 1)
                    subActoct = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9]+totsub[10])/((totsub[0]/12)*10)*100, 1)
                    subActnov = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9]+totsub[10]+totsub[11])/((totsub[0]/12)*11)*100, 1)
                    subActdic = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9]+totsub[10]+totsub[11]+totsub[12])/((totsub[0]/12)*12)*100, 1)


                totalActDetail = { 'cod_act': act[0], 'nombre_act': act[1],  'meta': totsub[0], 'ene': totsub[1], 'feb': totsub[2], 'mar': totsub[3], 'abr': totsub[4],
                                    'may': totsub[5], 'jun': totsub[6], 'jul': totsub[7], 'ago': totsub[8], 'set': totsub[9], 'oct': totsub[10],
                                    'nov': totsub[11], 'dic': totsub[12], 'avance': totsub[13], 'avporcent': avSubAct }

                totalAvActDetail = { 'cod_act': act[0], 'avAct_ene': subActene, 'avAct_feb': subActfeb, 'avAct_mar': subActmar, 'avAct_abr': subActabr, 'avAct_may': subActmay,
                                    'avAct_jun': subActjun, 'avAct_jul': subActjul, 'avAct_ago': subActago, 'avAct_set': subActset, 'avAct_oct': subActoct,
                                    'avAct_nov': subActnov, 'avAct_dic': subActdic }

                dataActDetail.append(totalActDetail)
                dataAvActDetail.append(totalAvActDetail)

            d.execute("""SELECT cod_act, cod_subact, meta into ESSALUD.dbo.metaprior_subact
                            from ESSALUD.dbo.metas_priorizadas where cod_centro=%s and anio=%s
                            group by cod_act, cod_subact, meta""" % (request.GET['eess'], request.GET['anio']))

            d.execute("""SELECT b.nombre nomsubact, c.meta, sum(IIF(ene is null, 0, ene)) ene, sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar,
                        sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may, sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul,
                        sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set], sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov,
                        sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+
                        sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))+sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+
                        sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0, [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                        FROM ESSALUD.dbo.avance a left join ESSALUD.dbo.subactividades b on a.cod_subact=b.codigo
                        left join ESSALUD.dbo.metaprior_subact c on a.cod_subact=c.cod_subact
                        where a.cod_centro=%s and a.anio=%s and a.cod_act=%s
                        group by b.nombre, c.meta
                        drop table ESSALUD.dbo.metaprior_subact""" % (request.GET['eess'], request.GET['anio'], "'"+act[0]+"'"))

            for dsub in d.fetchall():
                subAct = { 'cod_act': act[0], 'nombre_act': act[1],  'nombresub': dsub[0], 'meta': dsub[1], 'ene': dsub[2], 'feb': dsub[3],
                            'mar': dsub[4], 'abr': dsub[5], 'may': dsub[6], 'jun': dsub[7], 'jul': dsub[8], 'ago': dsub[9], 'set': dsub[10],
                            'oct': dsub[11], 'nov': dsub[12], 'dic': dsub[13], 'avance': dsub[14] }#'avporcent': round((dsub[13]/dsub[0])*100, 1)

                dataSubActiv.append(subAct)

        f = connection.cursor()
        f.execute("""select a.cod_centro, a.cod_act, b.nombre from ESSALUD.dbo.avance a left join ESSALUD.dbo.actividades b on a.cod_act=b.codigo
                    where cod_centro=%s and anio=%s group by a.cod_centro, a.cod_act, b.nombre""" % (request.GET['eess'], request.GET['anio']))

        nameActividad = []
        for nameAct in f.fetchall():
            names = { 'cod_act': nameAct[1], 'nombre': nameAct[2] }
            nameActividad.append(names)

        dataList.extend([resulTotal])
        dataList.extend([total])
        dataList.extend([totalAv])
        dataList.extend([dataAct])
        dataList.extend([dataActDetail])
        dataList.extend([dataAvActDetail])
        dataList.extend([dataSubActiv])
        dataList.extend([nameActividad])

        return HttpResponse(json.dumps(dataList), content_type='application/json')


class AdvMetasPriorXAct(View):
    def get(self, request, *args, **kwargs):
        e = connection.cursor()
        e.execute("""select (select sum(meta) from metas_priorizadas where cod_centro=%s and anio=%s and cod_act=%s) meta, sum(IIF(ene is null, 0, ene)) ene,
                        sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar, sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may,
                        sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul, sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set],
                        sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov, sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+
                        sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))
                        +sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0,
                        [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                        from avance where cod_centro=%s and anio=%s and cod_act=%s""" % (request.GET['eess'], request.GET['anio'], "'"+request.GET['act']+"'", request.GET['eess'], request.GET['anio'], "'"+request.GET['act']+"'"))

        for adv in e.fetchall():
            avance = { 'meta': adv[0], 'avAct_ene': adv[1], 'avAct_feb': adv[2], 'avAct_mar': adv[3], 'avAct_abr': adv[4], 'avAct_may': adv[5], 'avAct_jun': adv[6],
                        'avAct_jul': adv[7], 'avAct_ago': adv[8], 'avAct_set': adv[9], 'avAct_oct': adv[10], 'avAct_nov': adv[11], 'avAct_dic': adv[12] }

        return HttpResponse(json.dumps(avance), content_type='application/json')


class PrintGoals(View):
    def get(self, request, *args, **kwargs):
        locale.setlocale(locale.LC_ALL, "C")

        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "B2:Q2", "medium", "1F497D")

        ws.merge_cells('B2:Q2')
        ws.row_dimensions[2].height = 33
        ws.row_dimensions[6].height = 25

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 34
        ws.column_dimensions['C'].width = 10

        nameEess = Establecimiento.objects.get(codigo=request.GET['eess'])

        ws['B2'].font = Font(name='Aptos Narrow', size=20, bold=True, color='1F497D')
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B2'] = 'ESSALUD PASCO: '+ nameEess.nombre + ' - ' + str(request.GET['anio'])

        # ws.merge_cells('B4:B5')
        ws['B4'] = 'ACTIVIDADES'
        ws['B4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['C4'] = 'META'
        ws['C4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['D4'] = 'ENE'
        ws['D4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['E4'] = 'FEB'
        ws['E4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['F4'] = 'MAR'
        ws['F4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['G4'] = 'ABR'
        ws['G4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['G4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['H4'] = 'MAY'
        ws['H4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['H4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['H4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['I4'] = 'JUN'
        ws['I4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['I4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['I4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['J4'] = 'JUL'
        ws['J4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['J4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['J4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['K4'] = 'AGO'
        ws['K4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['K4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['K4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['L4'] = 'SET'
        ws['L4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['L4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['L4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['M4'] = 'OCT'
        ws['M4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['M4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['M4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['N4'] = 'NOV'
        ws['N4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['N4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['N4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['O4'] = 'DIC'
        ws['O4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['O4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['O4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['P4'] = 'AVANCE'
        ws['P4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['P4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['P4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['Q4'] = '% AVANCE AÑO'
        ws['Q4'].font = Font(name='Aptos Narrow', size=10, bold=True, color="FFFFFF")
        ws['Q4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Q4'].fill = PatternFill(start_color='2F4C70', end_color='2F4C70', fill_type='solid')
        ws['Q4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['B6'] = '% AVANCE AL MES'
        ws['B6'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B6'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B6'].fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
        ws['B6'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        a = connection.cursor()
        a.execute("""select (select sum(meta) from metas_priorizadas where cod_centro=%s and anio=%s) meta, sum(IIF(ene is null, 0, ene)) ene,
                    sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar, sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may,
                    sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul, sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set],
                    sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov, sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+
                    sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))
                    +sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0,
                    [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                    from ESSALUD.dbo.avance where cod_centro=%s and anio=%s""" % (request.GET['eess'], request.GET['anio'], request.GET['eess'], request.GET['anio']))

        for tot in a.fetchall():
            if tot[0] == 0 or tot[0] == None or tot[13] == None:
                avGeneralAnio = 0
                avEne = 0
                avFeb = 0
                avMar = 0
                avAbr = 0
                avMay = 0
                avJun = 0
                avJul = 0
                avAgo = 0
                avSet = 0
                avOct = 0
                avNov = 0
                avDic = 0
            else:
                avGeneralAnio = round((tot[13]/tot[0])*100, 1)
                avEne = round((tot[1]/(tot[0]/12))*100, 1)
                avFeb = round((tot[1]+tot[2])/((tot[0]/12)*2)*100, 1)
                avMar = round((tot[1]+tot[2]+tot[3])/((tot[0]/12)*3)*100, 1)
                avAbr = round((tot[1]+tot[2]+tot[3]+tot[4])/((tot[0]/12)*4)*100, 1)
                avMay = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5])/((tot[0]/12)*5)*100, 1)
                avJun = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6])/((tot[0]/12)*6)*100, 1)
                avJul = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7])/((tot[0]/12)*7)*100, 1)
                avAgo = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8])/((tot[0]/12)*8)*100, 1)
                avSet = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9])/((tot[0]/12)*9)*100, 1)
                avOct = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9]+tot[10])/((tot[0]/12)*10)*100, 1)
                avNov = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9]+tot[10]+tot[11])/((tot[0]/12)*11)*100, 1)
                avDic = round((tot[1]+tot[2]+tot[3]+tot[4]+tot[5]+tot[6]+tot[7]+tot[8]+tot[9]+tot[10]+tot[11]+tot[12])/((tot[0]/12)*12)*100, 1)

            ws.cell(row=5, column=3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=3).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=3).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=3).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=3).value = tot[0]

            ws.cell(row=5, column=4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=4).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=4).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=4).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=4).value = tot[1]

            ws.cell(row=5, column=5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=5).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=5).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=5).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=5).value = tot[2]

            ws.cell(row=5, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=6).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=6).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=6).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=6).value = tot[3]

            ws.cell(row=5, column=7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=7).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=7).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=7).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=7).value = tot[4]

            ws.cell(row=5, column=8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=8).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=8).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=8).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=8).value = tot[5]

            ws.cell(row=5, column=9).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=9).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=9).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=9).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=9).value = tot[6]

            ws.cell(row=5, column=10).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=10).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=10).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=10).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=10).value = tot[7]

            ws.cell(row=5, column=11).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=11).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=11).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=11).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=11).value = tot[8]

            ws.cell(row=5, column=12).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=12).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=12).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=12).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=12).value = tot[9]

            ws.cell(row=5, column=13).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=13).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=13).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=13).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=13).value = tot[10]

            ws.cell(row=5, column=14).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=14).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=14).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=14).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=14).value = tot[11]

            ws.cell(row=5, column=15).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=15).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=15).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=15).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=15).value = tot[12]

            ws.cell(row=5, column=16).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=16).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=16).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=16).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=16).value = tot[13]

            ws.cell(row=5, column=17).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=5, column=17).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=5, column=17).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=5, column=17).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ws.cell(row=5, column=17).value = str(avGeneralAnio)+' %'

            # -------
            traffic = '⬤'

            ws.cell(row=6, column=4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=4).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avEne < 81:
                ws.cell(row=6, column=4).font = Font(name='Calibri', size=10, color='FF2929')
            elif avEne > 80 and avEne < 96:
                ws.cell(row=6, column=4).font = Font(name='Calibri', size=10, color='FF9900')
            elif avEne > 95:
                ws.cell(row=6, column=4).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=4).value = str(avEne)+' % '+traffic

            ws.cell(row=6, column=5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=5).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avFeb < 81:
                ws.cell(row=6, column=5).font = Font(name='Calibri', size=10, color='FF2929')
            elif avFeb > 80 and avFeb < 96:
                ws.cell(row=6, column=5).font = Font(name='Calibri', size=10, color='FF9900')
            elif avFeb > 95:
                ws.cell(row=6, column=5).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=5).value = str(avFeb)+' % '+traffic

            ws.cell(row=6, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=6).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avMar < 81:
                ws.cell(row=6, column=6).font = Font(name='Calibri', size=10, color='FF2929')
            elif avMar > 80 and avMar < 96:
                ws.cell(row=6, column=6).font = Font(name='Calibri', size=10, color='FF9900')
            elif avMar > 95:
                ws.cell(row=6, column=6).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=6).value = str(avMar)+' % '+traffic

            ws.cell(row=6, column=7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=7).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avAbr < 81:
                ws.cell(row=6, column=7).font = Font(name='Calibri', size=10, color='FF2929')
            elif avAbr > 80 and avAbr < 96:
                ws.cell(row=6, column=7).font = Font(name='Calibri', size=10, color='FF9900')
            elif avAbr > 95:
                ws.cell(row=6, column=7).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=7).value = str(avAbr)+' % '+traffic

            ws.cell(row=6, column=8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=8).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avMay < 81:
                ws.cell(row=6, column=8).font = Font(name='Calibri', size=10, color='FF2929')
            elif avMay > 80 and avMay < 96:
                ws.cell(row=6, column=8).font = Font(name='Calibri', size=10, color='FF9900')
            elif avMay > 95:
                ws.cell(row=6, column=8).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=8).value = str(avMay)+' % '+traffic

            ws.cell(row=6, column=9).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=9).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avJun < 81:
                ws.cell(row=6, column=9).font = Font(name='Calibri', size=10, color='FF2929')
            elif avJun > 80 and avJun < 96:
                ws.cell(row=6, column=9).font = Font(name='Calibri', size=10, color='FF9900')
            elif avJun > 95:
                ws.cell(row=6, column=9).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=9).value = str(avJun)+' % '+traffic

            ws.cell(row=6, column=10).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=10).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avJul < 81:
                ws.cell(row=6, column=10).font = Font(name='Calibri', size=10, color='FF2929')
            elif avJul > 80 and avJul < 96:
                ws.cell(row=6, column=10).font = Font(name='Calibri', size=10, color='FF9900')
            elif avJul > 95:
                ws.cell(row=6, column=10).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=10).value = str(avJul)+' % '+traffic

            ws.cell(row=6, column=11).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=11).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avAgo < 81:
                ws.cell(row=6, column=11).font = Font(name='Calibri', size=10, color='FF2929')
            elif avAgo > 80 and avAgo < 96:
                ws.cell(row=6, column=11).font = Font(name='Calibri', size=10, color='FF9900')
            elif avAgo > 95:
                ws.cell(row=6, column=11).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=11).value = str(avAgo)+' % '+traffic

            ws.cell(row=6, column=12).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=12).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avSet < 81:
                ws.cell(row=6, column=12).font = Font(name='Calibri', size=10, color='FF2929')
            elif avSet > 80 and avSet < 96:
                ws.cell(row=6, column=12).font = Font(name='Calibri', size=10, color='FF9900')
            elif avSet > 95:
                ws.cell(row=6, column=12).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=12).value = str(avSet)+' % '+traffic

            ws.cell(row=6, column=13).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=13).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avOct < 81:
                ws.cell(row=6, column=13).font = Font(name='Calibri', size=10, color='FF2929')
            elif avOct > 80 and avOct < 96:
                ws.cell(row=6, column=13).font = Font(name='Calibri', size=10, color='FF9900')
            elif avOct > 95:
                ws.cell(row=6, column=13).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=13).value = str(avOct)+' % '+traffic

            ws.cell(row=6, column=14).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=14).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avNov < 81:
                ws.cell(row=6, column=14).font = Font(name='Calibri', size=10, color='FF2929')
            elif avNov > 80 and avNov < 96:
                ws.cell(row=6, column=14).font = Font(name='Calibri', size=10, color='FF9900')
            elif avNov > 95:
                ws.cell(row=6, column=14).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=14).value = str(avNov)+' % '+traffic

            ws.cell(row=6, column=15).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=6, column=15).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            if avDic < 81:
                ws.cell(row=6, column=15).font = Font(name='Calibri', size=10, color='FF2929')
            elif avDic > 80 and avDic < 96:
                ws.cell(row=6, column=15).font = Font(name='Calibri', size=10, color='FF9900')
            elif avDic > 95:
                ws.cell(row=6, column=15).font = Font(name='Calibri', size=10, color='00B050')
            ws.cell(row=6, column=15).value = str(avDic)+' % '+traffic

        b = connection.cursor()
        b.execute("""SELECT cod_act, sum(meta) meta into ESSALUD.dbo.metas_prior from metas_priorizadas where cod_centro=%s and anio=%s group by cod_act""" % (request.GET['eess'], request.GET['anio']))
        b.execute("""SELECT b.nombre, c.meta, sum(IIF(ene is null, 0, ene)) ene, sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar,
                    sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may, sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul,
                    sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set], sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov,
                    sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+
                    sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))+sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+
                    sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0, [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                    FROM ESSALUD.dbo.avance a left join ESSALUD.dbo.actividades b on a.cod_Act=b.codigo left join ESSALUD.dbo.metas_prior c on a.cod_act=c.cod_act
                    where a.cod_centro=%s and a.anio=%s group by b.nombre, c.meta
                    drop table ESSALUD.dbo.metas_prior""" % (request.GET['eess'], request.GET['anio']))

        cont = 7
        for avAct in b.fetchall():
            if avAct[1] == 0:
                avAcAnio = 0
            else:
                avAcAnio = round((avAct[14]/avAct[1])*100, 1)

            ws.cell(row=cont, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=2).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=2).value = avAct[0]

            ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=3).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=3).value = avAct[1]

            ws.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=4).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=4).value = avAct[2]

            ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=5).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=5).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=5).value = avAct[3]

            ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=6).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=6).value = avAct[4]

            ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=7).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=7).value = avAct[5]

            ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=8).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=8).value = avAct[6]

            ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=9).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=9).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=9).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=9).value = avAct[7]

            ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=10).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=10).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=10).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=10).value = avAct[8]

            ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=11).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=11).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=11).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=11).value = avAct[9]

            ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=12).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=12).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=12).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=12).value = avAct[10]

            ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=13).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=13).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=13).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=13).value = avAct[11]

            ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=14).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=14).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=14).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=14).value = avAct[12]

            ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=15).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=15).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=15).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=15).value = avAct[13]

            ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=16).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=16).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=16).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=16).value = avAct[14]

            ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=17).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=17).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=17).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row=cont, column=17).value = str(avAcAnio)+' %'

            cont = cont+1

        c = connection.cursor()
        c.execute("""SELECT a.cod_act, b.nombre from metas_priorizadas a left join actividades b on a.cod_act=b.codigo where cod_centro=%s and anio=%s
                        group by a.cod_act, b.nombre""" % (request.GET['eess'], request.GET['anio']))

        d = connection.cursor()
        e = connection.cursor()

        contAct = cont+1
        for act in c.fetchall():
            ws.cell(row=contAct, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=contAct, column=2).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
            ws.cell(row=contAct, column=2).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=contAct, column=2).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            ws.cell(row=contAct, column=2).value = 'ATENCIÓN '+act[1]

            ws.row_dimensions[contAct].height = 25
            ws.row_dimensions[contAct+1].height = 25

            ws.cell(row=contAct+1, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=contAct+1, column=2).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
            ws.cell(row=contAct+1, column=2).font = Font(name='Calibri', size=9, bold=True)
            ws.cell(row=contAct+1, column=2).fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
            ws.cell(row=contAct+1, column=2).value = '% AVANCE AL MES'

            e.execute("""select (select sum(meta) from metas_priorizadas where cod_centro=%s and anio=%s and cod_act=%s) meta, sum(IIF(ene is null, 0, ene)) ene,
                        sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar, sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may,
                        sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul, sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set],
                        sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov, sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+
                        sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))
                        +sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0,
                        [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                        from avance where cod_centro=%s and anio=%s and cod_act=%s""" % (request.GET['eess'], request.GET['anio'], "'"+act[0]+"'", request.GET['eess'], request.GET['anio'], "'"+act[0]+"'"))

            for totsub in e.fetchall():
                if totsub[0] == 0 or totsub[0] == None or totsub[13] == None:
                    avSubAct = 0
                else:
                    avSubAct = round((totsub[13]/totsub[0])*100, 1)

                ws.cell(row=contAct, column=3).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=3).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=3).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=3).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=3).value = totsub[0]

                ws.cell(row=contAct, column=4).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=4).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=4).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=4).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=4).value = totsub[1]

                ws.cell(row=contAct, column=5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=5).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=5).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=5).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=5).value = totsub[2]

                ws.cell(row=contAct, column=6).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=6).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=6).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=6).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=6).value = totsub[3]

                ws.cell(row=contAct, column=7).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=7).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=7).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=7).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=7).value = totsub[4]

                ws.cell(row=contAct, column=8).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=8).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=8).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=8).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=8).value = totsub[5]

                ws.cell(row=contAct, column=9).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=9).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=9).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=9).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=9).value = totsub[6]

                ws.cell(row=contAct, column=10).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=10).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=10).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=10).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=10).value = totsub[7]

                ws.cell(row=contAct, column=11).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=11).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=11).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=11).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=11).value = totsub[8]

                ws.cell(row=contAct, column=12).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=12).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=12).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=12).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=12).value = totsub[9]

                ws.cell(row=contAct, column=13).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=13).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=13).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=13).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=13).value = totsub[10]

                ws.cell(row=contAct, column=14).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=14).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=14).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=14).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=14).value = totsub[11]

                ws.cell(row=contAct, column=15).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=15).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=15).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=15).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=15).value = totsub[12]

                ws.cell(row=contAct, column=16).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=16).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=16).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=16).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=16).value = totsub[13]

                ws.cell(row=contAct, column=17).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct, column=17).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=17).font = Font(name='Calibri', size=9, bold=True)
                ws.cell(row=contAct, column=17).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                ws.cell(row=contAct, column=17).value = str(avSubAct)+' %'

                if totsub[0] == 0 or totsub[0] == None:
                    subActene = 0
                    subActfeb = 0
                    subActmar = 0
                    subActabr = 0
                    subActmay = 0
                    subActjun = 0
                    subActjul = 0
                    subActago = 0
                    subActset = 0
                    subActoct = 0
                    subActnov = 0
                    subActdic = 0
                else:
                    subActene = round((0 if totsub[1] == None else totsub[1]/(totsub[0]/12))*100, 1)
                    subActfeb = round((0 if totsub[1] == None else totsub[1]+totsub[2])/((totsub[0]/12)*2)*100, 1)
                    subActmar = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3])/((totsub[0]/12)*3)*100, 1)
                    subActabr = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4])/((totsub[0]/12)*4)*100, 1)
                    subActmay = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5])/((totsub[0]/12)*5)*100, 1)
                    subActjun = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6])/((totsub[0]/12)*6)*100, 1)
                    subActjul = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7])/((totsub[0]/12)*7)*100, 1)
                    subActago = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8])/((totsub[0]/12)*8)*100, 1)
                    subActset = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9])/((totsub[0]/12)*9)*100, 1)
                    subActoct = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9]+totsub[10])/((totsub[0]/12)*10)*100, 1)
                    subActnov = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9]+totsub[10]+totsub[11])/((totsub[0]/12)*11)*100, 1)
                    subActdic = round((0 if totsub[1] == None else totsub[1]+totsub[2]+totsub[3]+totsub[4]+totsub[5]+totsub[6]+totsub[7]+totsub[8]+totsub[9]+totsub[10]+totsub[11]+totsub[12])/((totsub[0]/12)*12)*100, 1)

                ws.cell(row=contAct+1, column=4).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=4).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActene < 81:
                    ws.cell(row=contAct+1, column=4).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActene > 80 and subActene < 96:
                    ws.cell(row=contAct+1, column=4).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActene > 95:
                    ws.cell(row=contAct+1, column=4).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=4).value = str(subActene)+' % '+traffic

                ws.cell(row=contAct+1, column=5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=5).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActfeb < 81:
                    ws.cell(row=contAct+1, column=5).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActfeb > 80 and subActfeb < 96:
                    ws.cell(row=contAct+1, column=5).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActfeb > 95:
                    ws.cell(row=contAct+1, column=5).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=5).value = str(subActfeb)+' % '+traffic

                ws.cell(row=contAct+1, column=6).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=6).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActmar < 81:
                    ws.cell(row=contAct+1, column=6).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActmar > 80 and subActmar < 96:
                    ws.cell(row=contAct+1, column=6).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActmar > 95:
                    ws.cell(row=contAct+1, column=6).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=6).value = str(subActmar)+' % '+traffic

                ws.cell(row=contAct+1, column=7).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=7).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActabr < 81:
                    ws.cell(row=contAct+1, column=7).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActabr > 80 and subActabr < 96:
                    ws.cell(row=contAct+1, column=7).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActabr > 95:
                    ws.cell(row=contAct+1, column=7).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=7).value = str(subActabr)+' % '+traffic

                ws.cell(row=contAct+1, column=8).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=8).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActmay < 81:
                    ws.cell(row=contAct+1, column=8).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActmay > 80 and subActmay < 96:
                    ws.cell(row=contAct+1, column=8).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActmay > 95:
                    ws.cell(row=contAct+1, column=8).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=8).value = str(subActmay)+' % '+traffic

                ws.cell(row=contAct+1, column=9).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=9).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActjun < 81:
                    ws.cell(row=contAct+1, column=9).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActjun > 80 and subActjun < 96:
                    ws.cell(row=contAct+1, column=9).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActjun > 95:
                    ws.cell(row=contAct+1, column=9).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=9).value = str(subActjun)+' % '+traffic

                ws.cell(row=contAct+1, column=10).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=10).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActjul < 81:
                    ws.cell(row=contAct+1, column=10).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActjul > 80 and subActjul < 96:
                    ws.cell(row=contAct+1, column=10).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActjul > 95:
                    ws.cell(row=contAct+1, column=10).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=10).value = str(subActjul)+' % '+traffic

                ws.cell(row=contAct+1, column=11).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=11).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActago < 81:
                    ws.cell(row=contAct+1, column=11).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActago > 80 and subActago < 96:
                    ws.cell(row=contAct+1, column=11).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActago > 95:
                    ws.cell(row=contAct+1, column=11).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=11).value = str(subActago)+' % '+traffic

                ws.cell(row=contAct+1, column=12).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=12).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActset < 81:
                    ws.cell(row=contAct+1, column=12).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActset > 80 and subActset < 96:
                    ws.cell(row=contAct+1, column=12).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActset > 95:
                    ws.cell(row=contAct+1, column=12).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=12).value = str(subActset)+' % '+traffic

                ws.cell(row=contAct+1, column=13).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=13).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActoct < 81:
                    ws.cell(row=contAct+1, column=13).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActoct > 80 and subActoct < 96:
                    ws.cell(row=contAct+1, column=13).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActoct > 95:
                    ws.cell(row=contAct+1, column=13).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=13).value = str(subActoct)+' % '+traffic

                ws.cell(row=contAct+1, column=14).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=14).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActnov < 81:
                    ws.cell(row=contAct+1, column=14).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActnov > 80 and subActnov < 96:
                    ws.cell(row=contAct+1, column=14).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActnov > 95:
                    ws.cell(row=contAct+1, column=14).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=14).value = str(subActnov)+' % '+traffic

                ws.cell(row=contAct+1, column=15).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=contAct+1, column=15).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                if subActdic < 81:
                    ws.cell(row=contAct+1, column=15).font = Font(name='Calibri', size=10, color='FF2929')
                elif subActdic > 80 and subActdic < 96:
                    ws.cell(row=contAct+1, column=15).font = Font(name='Calibri', size=10, color='FF9900')
                elif subActdic > 95:
                    ws.cell(row=contAct+1, column=15).font = Font(name='Calibri', size=10, color='00B050')
                ws.cell(row=contAct+1, column=15).value = str(subActdic)+' % '+traffic

            contAct = contAct+2

            d.execute("""SELECT cod_act, cod_subact, meta into ESSALUD.dbo.metaprior_subact
                            from ESSALUD.dbo.metas_priorizadas where cod_centro=%s and anio=%s
                            group by cod_act, cod_subact, meta""" % (request.GET['eess'], request.GET['anio']))

            d.execute("""SELECT b.nombre, c.meta, sum(IIF(ene is null, 0, ene)) ene, sum(IIF(feb is null, 0, feb)) feb, sum(IIF(mar is null, 0, mar)) mar,
                        sum(IIF(abr is null, 0, abr)) abr, sum(IIF(may is null, 0, may)) may, sum(IIF(jun is null, 0, jun)) jun, sum(IIF(jul is null, 0, jul)) jul,
                        sum(IIF(ago is null, 0, ago)) ago, sum(IIF([set] is null, 0, [set])) [set], sum(IIF(oct is null, 0, oct)) oct, sum(IIF(nov is null, 0, nov)) nov,
                        sum(IIF([dic] is null, 0, [dic])) dic, (sum(IIF(ene is null, 0, ene))+sum(IIF(feb is null, 0, feb))+sum(IIF(mar is null, 0, mar))+
                        sum(IIF(abr is null, 0, abr))+sum(IIF(may is null, 0, may))+sum(IIF(jun is null, 0, jun))+sum(IIF(jul is null, 0, jul))+sum(IIF(ago is null, 0, ago))+
                        sum(IIF([set] is null, 0, [set]))+sum(IIF([oct] is null, 0, [oct]))+sum(IIF([nov] is null, 0, [nov]))+sum(IIF([dic] is null, 0, [dic]))) avance
                        FROM ESSALUD.dbo.avance a left join ESSALUD.dbo.subactividades b on a.cod_subact=b.codigo
                        left join ESSALUD.dbo.metaprior_subact c on a.cod_subact=c.cod_subact
                        where a.cod_centro=%s and a.anio=%s and a.cod_act=%s
                        group by b.nombre, c.meta
                        drop table ESSALUD.dbo.metaprior_subact""" % (request.GET['eess'], request.GET['anio'], "'"+act[0]+"'"))

            for sub in d.fetchall():
                if sub[1] == 0 or sub[1] == None:
                    avSubAct2 = 0
                else:
                    avSubAct2 = round((sub[14]/sub[1])*100, 1)

                ws.cell(row=contAct, column=2).alignment = Alignment(horizontal="right")
                ws.cell(row=contAct, column=2).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=2).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=2).value = sub[0]

                ws.cell(row=contAct, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=3).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=3).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=3).value = sub[1]

                ws.cell(row=contAct, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=4).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=4).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=4).value = sub[2]

                ws.cell(row=contAct, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=5).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=5).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=5).value = sub[3]

                ws.cell(row=contAct, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=6).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=6).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=6).value = sub[4]

                ws.cell(row=contAct, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=7).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=7).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=7).value = sub[5]

                ws.cell(row=contAct, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=8).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=8).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=8).value = sub[6]

                ws.cell(row=contAct, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=9).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=9).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=9).value = sub[7]

                ws.cell(row=contAct, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=10).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=10).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=10).value = sub[8]

                ws.cell(row=contAct, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=11).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=11).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=11).value = sub[9]

                ws.cell(row=contAct, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=12).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=12).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=12).value = sub[10]

                ws.cell(row=contAct, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=13).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=13).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=13).value = sub[11]

                ws.cell(row=contAct, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=14).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=14).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=14).value = sub[12]

                ws.cell(row=contAct, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=15).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=15).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=15).value = sub[13]

                ws.cell(row=contAct, column=16).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=16).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=16).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=16).value = sub[14]

                ws.cell(row=contAct, column=17).alignment = Alignment(horizontal="center")
                ws.cell(row=contAct, column=17).border = Border(left=Side(border_style="medium", color="4BACC6"), right=Side(border_style="medium", color="4BACC6"), top=Side(border_style="medium", color="4BACC6"), bottom=Side(border_style="medium", color="4BACC6"))
                ws.cell(row=contAct, column=17).font = Font(name='Calibri', size=9)
                ws.cell(row=contAct, column=17).value = str(avSubAct2) + ' %'

                contAct = contAct+1

            contAct = contAct+1


        nombre_archivo = "ESSALUD - SEGUIMIENTO DE METAS PRIORIZADAS.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL NIÑOS RN'
        wb.save(response)
        return response


class InmunizationView(TemplateView):
    template_name = 'inmunizations/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.session['sytem']['typeca'] == 'CA':
            context['establecimiento'] = Establecimiento.objects.filter(codigo=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DS':
            context['establecimiento'] = Establecimiento.objects.filter(dist_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'PR':
            context['establecimiento'] = Establecimiento.objects.filter(prov_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DP':
            context['establecimiento'] = Establecimiento.objects.filter(dep_id=self.request.session['sytem']['codeca'])
        return context


class NominalInmunization(View):
    def get(self, request, *args, **kwargs):
        if request.GET['edad'] == '2':
            medad = [2,3]
        elif request.GET['edad'] == '5':
            medad = [9,10,11,12,13,14,15,16,17]
        elif request.GET['edad'] == '6':
            medad = [12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,
                     41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59]
        elif request.GET['edad'] == '7':
            medad = [60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100]
        else:
            medad = request.GET['edad']

        if request.GET['eess'] == 'TODOS':
            if request.GET['tipo'] == 'MINSA':
                if self.request.session['sytem']['typeca'] == 'CA':
                    nominal = InmunizationMinsa.objects.filter(cod_eess=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DS':
                    nominal = InmunizationMinsa.objects.filter(cod_dist=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'PR':
                    nominal = InmunizationMinsa.objects.filter(cod_prov=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DP':
                    nominal = InmunizationMinsa.objects.filter(cod_dep=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')

            elif request.GET['tipo'] == 'ESSALUD':
                if self.request.session['sytem']['typeca'] == 'CA':
                    nominal = Inmunization.objects.filter(cod_eess=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DS':
                    nominal = Inmunization.objects.filter(cod_dist=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'PR':
                    nominal = Inmunization.objects.filter(cod_prov=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
                elif self.request.session['sytem']['typeca'] == 'DP':
                    nominal = Inmunization.objects.filter(cod_dep=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
        else:
            if request.GET['tipo'] == 'MINSA':
                nominal = InmunizationMinsa.objects.filter(cod_eess=request.GET['eess'], edad__in=medad).order_by('cod_eess')
            elif request.GET['tipo'] == 'ESSALUD':
                nominal = Inmunization.objects.filter(cod_eess=request.GET['eess'], edad__in=medad).order_by('cod_eess')

        nominal = serializers.serialize('json', nominal, indent=2, use_natural_foreign_keys=True)

        return HttpResponse(nominal, content_type='application/json')


class ConteoInmunization(View):
    def get(self, request, *args, **kwargs):
        if request.GET['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                nominal = Inmunization_C.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'DS':
                nominal = Inmunization_C.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'PR':
                nominal = Inmunization_C.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'DP':
                nominal = Inmunization_C.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')
        else:
            nominal = Inmunization_C.objects.filter(cod_eess=request.GET['eess'], anio=request.GET['anio'], mes=request.GET['mes']).order_by('cod_eess')

        nominal = serializers.serialize('json', nominal, indent=2, use_natural_foreign_keys=True)

        return HttpResponse(nominal, content_type='application/json')


class PrintInmunization(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        if request.GET['edad'] == '2':
            medad = [2,3]
        elif request.GET['edad'] == '5':
            medad = [9,10,11,12,13,14,15,16,17]
        elif request.GET['edad'] == '6':
            medad = [12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,
                     41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59]
        elif request.GET['edad'] == '7':
            medad = [60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,
                     99,100]
        else:
            medad = request.GET['edad']

        if request.GET['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                totNominal = Inmunization.objects.filter(cod_eess=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'DS':
                totNominal = Inmunization.objects.filter(cod_dist=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'PR':
                totNominal = Inmunization.objects.filter(cod_prov=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'DP':
                totNominal = Inmunization.objects.filter(cod_dep=self.request.session['sytem']['codeca'], edad__in=medad).order_by('cod_eess')

        elif request.GET['eess'] != 'TODOS':
            totNominal = Inmunization.objects.filter(cod_eess=request.GET['eess'], edad__in=medad).order_by('cod_eess')

        totNominal = json.loads(serializers.serialize('json', totNominal, indent=2, use_natural_foreign_keys=True))

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws['A8'] = '#'
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['A8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['B8'] = 'Centro Asistencial'
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['B8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['C8'] = 'Documento'
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['C8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['D8'] = 'Apellidos y Nombres'
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['D8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['E8'] = 'Fecha Nacido'
        ws['E8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['E8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        ws['F8'] = 'Edad'
        ws['F8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws['F8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

        if request.GET['edad'] == '0':
            set_border(self, ws, "A2:T2", "medium", "57267C")
            set_border(self, ws, "A4:T4", "medium", "366092")
            set_border(self, ws, "A6:T6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 12
            ws.column_dimensions['N'].width = 12
            ws.column_dimensions['O'].width = 12
            ws.column_dimensions['P'].width = 12
            ws.column_dimensions['Q'].width = 12
            ws.column_dimensions['R'].width = 12
            ws.column_dimensions['S'].width = 12
            ws.column_dimensions['T'].width = 12

            ws.merge_cells('B2:T2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES RECIEN NACIDO Y MENORES DE 1 AÑO'

            ws.merge_cells('A4:T4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:T6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'BCG'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'HVB'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['I8'] = 'IPV 2M'
            ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['J8'] = 'ROTA 2M'
            ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['K8'] = 'PENTA 2M'
            ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['L8'] = 'NEUMO 2M'
            ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['M8'] = 'ROTA 4M'
            ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['M8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['N8'] = 'IPV 4M'
            ws['N8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['N8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['N8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['N8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['O8'] = 'PENTA 4M'
            ws['O8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['O8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['O8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['O8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['P8'] = 'NEUMO 4M'
            ws['P8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['P8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['P8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['P8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['Q8'] = 'IPV 6M'
            ws['Q8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['Q8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['Q8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['Q8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['R8'] = 'PENTA 6M'
            ws['R8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['R8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['R8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['R8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['S8'] = 'INFL 6M'
            ws['S8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['S8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['S8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['S8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['T8'] = 'INFL 7M'
            ws['T8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['T8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['T8'].fill = PatternFill(start_color='dff0c7', end_color='dff0c7', fill_type='solid')
            ws['T8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['bcg']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['hvb']

                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = nom['fields']['apo2m']

                    ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = nom['fields']['rota2m']

                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=11).value = nom['fields']['penta2m']

                    ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=12).value = nom['fields']['neumo2m']

                    ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=13).value = nom['fields']['rota4m']

                    ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=14).value = nom['fields']['apo4m']

                    ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=15).value = nom['fields']['penta4m']

                    ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=16).value = nom['fields']['neumo4m']

                    ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=17).value = nom['fields']['apo6m']

                    ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=18).value = nom['fields']['penta6m']

                    ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=19).value = nom['fields']['infl6m']

                    ws.cell(row=cont, column=20).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=20).value = nom['fields']['infl7m']

                    cont = cont+1
                    num = num+1

        elif request.GET['edad'] == '1':
            set_border(self, ws, "A2:Q2", "medium", "57267C")
            set_border(self, ws, "A4:Q4", "medium", "366092")
            set_border(self, ws, "A6:Q6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 12
            ws.column_dimensions['N'].width = 12
            ws.column_dimensions['O'].width = 12
            ws.column_dimensions['P'].width = 12
            ws.column_dimensions['Q'].width = 12

            ws.merge_cells('B2:Q2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES 1 AÑO'

            ws.merge_cells('A4:Q4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:Q6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'SPR1'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'NEUMO 12M'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['I8'] = 'VARICELA'
            ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['J8'] = 'INFL 1A'
            ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['K8'] = 'INFL B'
            ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['L8'] = 'HIB'
            ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['M8'] = 'AMA'
            ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['M8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['N8'] = 'SPR2'
            ws['N8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['N8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['N8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['N8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['O8'] = 'REF DPT'
            ws['O8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['O8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['O8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['O8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['P8'] = 'REF IPV'
            ws['P8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['P8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['P8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['P8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['Q8'] = 'HEPATITIS'
            ws['Q8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['Q8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['Q8'].fill = PatternFill(start_color='dfd2f4', end_color='dfd2f4', fill_type='solid')
            ws['Q8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['spr1']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['neumo6m']

                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = nom['fields']['varicela']

                    ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = nom['fields']['infl1a']

                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=11).value = nom['fields']['inflb']

                    ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=12).value = nom['fields']['hib']

                    ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=13).value = nom['fields']['ama']

                    ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=14).value = nom['fields']['spr2']

                    ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=15).value = nom['fields']['ref_dpt']

                    ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=16).value = nom['fields']['ref_ipv']

                    ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=17).value = nom['fields']['hepatitis']

                    cont = cont+1
                    num = num+1

        elif request.GET['edad'] == '2':
            set_border(self, ws, "A2:I2", "medium", "57267C")
            set_border(self, ws, "A4:I4", "medium", "366092")
            set_border(self, ws, "A6:I6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12

            ws.merge_cells('B2:I2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES 2 y 3 AÑOs'

            ws.merge_cells('A4:I4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:I6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'VARICELA'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='d6eaec', end_color='d6eaec', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'INFLUENZA 2A'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='d6eaec', end_color='d6eaec', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['I8'] = 'INFLUENZA 3A'
            ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I8'].fill = PatternFill(start_color='d6eaec', end_color='d6eaec', fill_type='solid')
            ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['varicela2']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['infl2a']

                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = nom['fields']['infl3a']

                    cont = cont+1
                    num = num+1

        elif request.GET['edad'] == '4':
            set_border(self, ws, "A2:I2", "medium", "57267C")
            set_border(self, ws, "A4:I4", "medium", "366092")
            set_border(self, ws, "A6:I6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12

            ws.merge_cells('B2:I2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES 4 AÑOS'

            ws.merge_cells('A4:I4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:I6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'INFLUENZA'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'REF APO'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['I8'] = 'REF DPT'
            ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I8'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')
            ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['infl4a']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['ref_apo2']

                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = nom['fields']['ref_dpt2']

                    cont = cont+1
                    num = num+1

        elif request.GET['edad'] == '5':
            set_border(self, ws, "A2:I2", "medium", "57267C")
            set_border(self, ws, "A4:I4", "medium", "366092")
            set_border(self, ws, "A6:I6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12

            ws.merge_cells('B2:I2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES ADOLESCENTES'

            ws.merge_cells('A4:I4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:I6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'VPH 1'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='d1d1d1', end_color='d1d1d1', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'VPH 2'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='d1d1d1', end_color='d1d1d1', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['I8'] = 'VPH DU'
            ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I8'].fill = PatternFill(start_color='d1d1d1', end_color='d1d1d1', fill_type='solid')
            ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['vph1']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['vph2']

                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = nom['fields']['vph_du']

                    cont = cont+1
                    num = num+1

        elif request.GET['edad'] == '6':
            set_border(self, ws, "A2:M2", "medium", "57267C")
            set_border(self, ws, "A4:M4", "medium", "366092")
            set_border(self, ws, "A6:M6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 12

            ws.merge_cells('B2:M2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES GESTANTES'

            ws.merge_cells('A4:M4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:M6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'DT 1'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'DT 2'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['I8'] = 'DTPA'
            ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['J8'] = 'INFLUENZA'
            ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['K8'] = 'HEPATITISB 1'
            ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['L8'] = 'HEPATITISB 2'
            ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['M8'] = 'HEPATITISB 3'
            ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['M8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M8'].fill = PatternFill(start_color='e7b6d0', end_color='e7b6d0', fill_type='solid')
            ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['gest_dt1']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['gest_dt2']

                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = nom['fields']['gest_tdap']

                    ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = nom['fields']['gest_infl']

                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=11).value = nom['fields']['hepatitisb1']

                    ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=12).value = nom['fields']['hepatitisb2']

                    ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=13).value = nom['fields']['hepatitisb3']

                    cont = cont+1
                    num = num+1

        elif request.GET['edad'] == '7':
            set_border(self, ws, "A2:H2", "medium", "57267C")
            set_border(self, ws, "A4:H4", "medium", "366092")
            set_border(self, ws, "A6:H6", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 33
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12

            ws.merge_cells('B2:H2')
            ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'ESSALUD PASCO: SEGUIMIENTO NOMINAL DE INMUNIZACIONES ADULTO MAYOR'

            ws.merge_cells('A4:H4')
            ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['A4'] = 'CODIFICACION: '

            ws.merge_cells('A6:H6')
            ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['G8'] = 'INLUENZA'
            ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G8'].fill = PatternFill(start_color='f2f2c1', end_color='f2f2c1', fill_type='solid')
            ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            ws['H8'] = 'NEUMOCOCO'
            ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H8'].fill = PatternFill(start_color='f2f2c1', end_color='f2f2c1', fill_type='solid')
            ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))

            cont = 9
            cant = len(totNominal)
            num = 1
            if cant > 0:
                for nom in totNominal:
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=1).value = num

                    ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=2).value = nom['fields']['eess']

                    ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=3).value = nom['fields']['documento']

                    ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                    ws.cell(row=cont, column=4).value = nom['fields']['paciente']

                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=5).value = nom['fields']['fec_nac']

                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = nom['fields']['edad']

                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = nom['fields']['infl_adult']

                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = nom['fields']['neumo_adult']

                    cont = cont+1
                    num = num+1


        nombre_archivo = "DEIT_PASCO NOMINAL INMUNIZACIONES.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL NIÑOS'
        wb.save(response)
        return response
