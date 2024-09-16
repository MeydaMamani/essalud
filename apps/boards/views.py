from django.shortcuts import render
from django.views.generic import TemplateView, View
from django.http import JsonResponse, HttpResponse, QueryDict
from django.core import serializers

from apps.boards.models import Coverage, metaCoverage
from apps.main.models import Provincia, Distrito, Establecimiento

from django.db.models import Sum, F, FloatField
from django.db.models.functions import Cast
from django.db import connection
from datetime import date, datetime

import datetime
import json
import locale

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color


# Create your views here.
class CoverageView(TemplateView):
    template_name = 'coverage/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.session['sytem']['typeca'] == 'CA':
            context['establecimiento'] = Establecimiento.objects.filter(codigo=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DS':
            context['establecimiento'] = Establecimiento.objects.filter(dist_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'PR':
            context['establecimiento'] = Establecimiento.objects.filter(prov_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DP':
            context['establecimiento'] = Establecimiento.objects.filter(dep_id=self.request.session['sytem']['codeca'])
        return context


class ListCoverage(TemplateView):
    def get(self, request, *args, **kwargs):
        dataList = []
        today = date.today()
        if request.GET['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                metarn = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('rn'))
                metaLessyear = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('less_1year'))
                metaOneyear = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('one_year'))
                metaFouryear = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('four_year'))
                metaVphGirl = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('girl9_13'))
                metaVphBoy = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('boy9_13'))
                metaGest = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('pregnant'))
                metaInfAdult = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult60'))
                metaNeumoAdult = metaCoverage.objects.filter(anio=today.year, cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult30'))

                bcg = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBcg=Sum('bcg'), av_bcg=Sum('bcg', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                hvb = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHvb=Sum('hvb'), av_hvb=Sum('hvb', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                rota = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalRota=Sum('rota2'), av_rota=Sum('rota2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                apo = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo=Sum('apo3'), av_apo=Sum('apo3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                penta = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalPenta=Sum('penta3'), av_penta=Sum('penta3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                infl2 = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInfl2=Sum('infl2'), av_infl2=Sum('infl2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                neumo3 = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumo3=Sum('neumo3'), av_neumo3=Sum('neumo3', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                var = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalVar=Sum('varicela1'), av_var=Sum('varicela1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr1 = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr1=Sum('spr1'), av_spr1=Sum('spr1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                ama = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalAma=Sum('ama'), av_ama=Sum('ama', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                hav = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHav=Sum('hav'), av_hav=Sum('hav', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr2 = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr2=Sum('spr2'), av_spr2=Sum('spr2', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                dpt2_ref = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpt2=Sum('dpt2_ref'), av_dpt2=Sum('dpt2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                apo2_ref = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo2=Sum('apo2_ref'), av_apo2=Sum('apo2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                vph_girl = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalGirl=Sum('vph_girl'), av_girl=Sum('vph_girl', output_field=FloatField()) / Cast(metaVphGirl['total'], FloatField()) * 100)
                vph_boy = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBoy=Sum('vph_boy'), av_boy=Sum('vph_boy', output_field=FloatField()) / Cast(metaVphBoy['total'], FloatField()) * 100)
                gestante = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpta=Sum('dpta'), av_gest=Sum('dpta', output_field=FloatField()) / Cast(metaGest['total'], FloatField()) * 100)
                infAdult = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInflAdult=Sum('infl_adult'), av_infAdul=Sum('infl_adult', output_field=FloatField()) / Cast(metaInfAdult['total'], FloatField()) * 100)
                neumoAdult = Coverage.objects.filter(cod_eess=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumoAdult=Sum('neumo_adult'), av_NeumoAdul=Sum('neumo_adult', output_field=FloatField()) / Cast(metaNeumoAdult['total'], FloatField()) * 100)

            elif self.request.session['sytem']['typeca'] == 'DS':
                metarn = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('rn'))
                metaLessyear = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('less_1year'))
                metaOneyear = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('one_year'))
                metaFouryear = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('four_year'))
                metaVphGirl = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('girl9_13'))
                metaVphBoy = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('boy9_13'))
                metaGest = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('pregnant'))
                metaInfAdult = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult60'))
                metaNeumoAdult = metaCoverage.objects.filter(anio=today.year, cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult30'))

                bcg = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBcg=Sum('bcg'), av_bcg=Sum('bcg', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                hvb = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHvb=Sum('hvb'), av_hvb=Sum('hvb', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                rota = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalRota=Sum('rota2'), av_rota=Sum('rota2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                apo = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo=Sum('apo3'), av_apo=Sum('apo3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                penta = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalPenta=Sum('penta3'), av_penta=Sum('penta3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                infl2 = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInfl2=Sum('infl2'), av_infl2=Sum('infl2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                neumo3 = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumo3=Sum('neumo3'), av_neumo3=Sum('neumo3', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                var = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalVar=Sum('varicela1'), av_var=Sum('varicela1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr1 = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr1=Sum('spr1'), av_spr1=Sum('spr1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                ama = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalAma=Sum('ama'), av_ama=Sum('ama', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                hav = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHav=Sum('hav'), av_hav=Sum('hav', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr2 = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr2=Sum('spr2'), av_spr2=Sum('spr2', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                dpt2_ref = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpt2=Sum('dpt2_ref'), av_dpt2=Sum('dpt2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                apo2_ref = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo2=Sum('apo2_ref'), av_apo2=Sum('apo2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                vph_girl = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalGirl=Sum('vph_girl'), av_girl=Sum('vph_girl', output_field=FloatField()) / Cast(metaVphGirl['total'], FloatField()) * 100)
                vph_boy = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBoy=Sum('vph_boy'), av_boy=Sum('vph_boy', output_field=FloatField()) / Cast(metaVphBoy['total'], FloatField()) * 100)
                gestante = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpta=Sum('dpta'), av_gest=Sum('dpta', output_field=FloatField()) / Cast(metaGest['total'], FloatField()) * 100)
                infAdult = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInflAdult=Sum('infl_adult'), av_infAdul=Sum('infl_adult', output_field=FloatField()) / Cast(metaInfAdult['total'], FloatField()) * 100)
                neumoAdult = Coverage.objects.filter(cod_dist=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumoAdult=Sum('neumo_adult'), av_NeumoAdul=Sum('neumo_adult', output_field=FloatField()) / Cast(metaNeumoAdult['total'], FloatField()) * 100)

            elif self.request.session['sytem']['typeca'] == 'PR':
                metarn = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('rn'))
                metaLessyear = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('less_1year'))
                metaOneyear = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('one_year'))
                metaFouryear = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('four_year'))
                metaVphGirl = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('girl9_13'))
                metaVphBoy = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('boy9_13'))
                metaGest = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('pregnant'))
                metaInfAdult = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult60'))
                metaNeumoAdult = metaCoverage.objects.filter(anio=today.year, cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult30'))

                bcg = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBcg=Sum('bcg'), av_bcg=Sum('bcg', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                hvb = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHvb=Sum('hvb'), av_hvb=Sum('hvb', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                rota = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalRota=Sum('rota2'), av_rota=Sum('rota2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                apo = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo=Sum('apo3'), av_apo=Sum('apo3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                penta = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalPenta=Sum('penta3'), av_penta=Sum('penta3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                infl2 = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInfl2=Sum('infl2'), av_infl2=Sum('infl2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                neumo3 = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumo3=Sum('neumo3'), av_neumo3=Sum('neumo3', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                var = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalVar=Sum('varicela1'), av_var=Sum('varicela1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr1 = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr1=Sum('spr1'), av_spr1=Sum('spr1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                ama = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalAma=Sum('ama'), av_ama=Sum('ama', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                hav = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHav=Sum('hav'), av_hav=Sum('hav', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr2 = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr2=Sum('spr2'), av_spr2=Sum('spr2', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                dpt2_ref = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpt2=Sum('dpt2_ref'), av_dpt2=Sum('dpt2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                apo2_ref = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo2=Sum('apo2_ref'), av_apo2=Sum('apo2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                vph_girl = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalGirl=Sum('vph_girl'), av_girl=Sum('vph_girl', output_field=FloatField()) / Cast(metaVphGirl['total'], FloatField()) * 100)
                vph_boy = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBoy=Sum('vph_boy'), av_boy=Sum('vph_boy', output_field=FloatField()) / Cast(metaVphBoy['total'], FloatField()) * 100)
                gestante = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpta=Sum('dpta'), av_gest=Sum('dpta', output_field=FloatField()) / Cast(metaGest['total'], FloatField()) * 100)
                infAdult = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInflAdult=Sum('infl_adult'), av_infAdul=Sum('infl_adult', output_field=FloatField()) / Cast(metaInfAdult['total'], FloatField()) * 100)
                neumoAdult = Coverage.objects.filter(cod_prov=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumoAdult=Sum('neumo_adult'), av_NeumoAdul=Sum('neumo_adult', output_field=FloatField()) / Cast(metaNeumoAdult['total'], FloatField()) * 100)

            elif self.request.session['sytem']['typeca'] == 'DP':
                metarn = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('rn'))
                metaLessyear = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('less_1year'))
                metaOneyear = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('one_year'))
                metaFouryear = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('four_year'))
                metaVphGirl = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('girl9_13'))
                metaVphBoy = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('boy9_13'))
                metaGest = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('pregnant'))
                metaInfAdult = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult60'))
                metaNeumoAdult = metaCoverage.objects.filter(anio=today.year, cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('adult30'))

                bcg = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBcg=Sum('bcg'), av_bcg=Sum('bcg', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                hvb = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHvb=Sum('hvb'), av_hvb=Sum('hvb', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
                rota = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalRota=Sum('rota2'), av_rota=Sum('rota2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                apo = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo=Sum('apo3'), av_apo=Sum('apo3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                penta = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalPenta=Sum('penta3'), av_penta=Sum('penta3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                infl2 = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInfl2=Sum('infl2'), av_infl2=Sum('infl2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
                neumo3 = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumo3=Sum('neumo3'), av_neumo3=Sum('neumo3', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                var = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalVar=Sum('varicela1'), av_var=Sum('varicela1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr1 = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr1=Sum('spr1'), av_spr1=Sum('spr1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                ama = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalAma=Sum('ama'), av_ama=Sum('ama', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                hav = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalHav=Sum('hav'), av_hav=Sum('hav', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                spr2 = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalSpr2=Sum('spr2'), av_spr2=Sum('spr2', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
                dpt2_ref = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpt2=Sum('dpt2_ref'), av_dpt2=Sum('dpt2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                apo2_ref = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalApo2=Sum('apo2_ref'), av_apo2=Sum('apo2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
                vph_girl = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalGirl=Sum('vph_girl'), av_girl=Sum('vph_girl', output_field=FloatField()) / Cast(metaVphGirl['total'], FloatField()) * 100)
                vph_boy = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalBoy=Sum('vph_boy'), av_boy=Sum('vph_boy', output_field=FloatField()) / Cast(metaVphBoy['total'], FloatField()) * 100)
                gestante = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalDpta=Sum('dpta'), av_gest=Sum('dpta', output_field=FloatField()) / Cast(metaGest['total'], FloatField()) * 100)
                infAdult = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalInflAdult=Sum('infl_adult'), av_infAdul=Sum('infl_adult', output_field=FloatField()) / Cast(metaInfAdult['total'], FloatField()) * 100)
                neumoAdult = Coverage.objects.filter(cod_dep=self.request.session['sytem']['codeca'], anio=today.year, mes__range=[1, request.GET['month']]).aggregate(totalNeumoAdult=Sum('neumo_adult'), av_NeumoAdul=Sum('neumo_adult', output_field=FloatField()) / Cast(metaNeumoAdult['total'], FloatField()) * 100)

            resulTotal = {
                'av_bcg': bcg['av_bcg'], 'av_hvb': hvb['av_hvb'], 'av_rota': rota['av_rota'],
                'av_apo': apo['av_apo'], 'av_penta': penta['av_penta'], 'av_infl2': infl2['av_infl2'],
                'av_neumo3': neumo3['av_neumo3'], 'av_var': var['av_var'], 'av_spr1': spr1['av_spr1'],
                'av_ama': ama['av_ama'], 'av_hav': hav['av_hav'], 'av_spr2': spr2['av_spr2'], 'av_dpt2': dpt2_ref['av_dpt2'],
                'av_apo2': apo2_ref['av_apo2'], 'av_girl': vph_girl['av_girl'], 'av_boy': vph_boy['av_boy'],
                'av_gest': gestante['av_gest'], 'av_infAdul': infAdult['av_infAdul'], 'av_NeumoAdul': neumoAdult['av_NeumoAdul'],
            }

        else:
            metarn = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('rn'))
            metaLessyear = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('less_1year'))
            metaOneyear = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('one_year'))
            metaFouryear = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('four_year'))
            metaVphGirl = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('girl9_13'))
            metaVphBoy = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('boy9_13'))
            metaGest = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('pregnant'))
            metaInfAdult = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('adult60'))
            metaNeumoAdult = metaCoverage.objects.filter(anio=today.year, cod_eess=request.GET['eess']).aggregate(total=Sum('adult30'))

            bcg = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalBcg=Sum('bcg'), av_bcg=Sum('bcg', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
            hvb = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalHvb=Sum('hvb'), av_hvb=Sum('hvb', output_field=FloatField()) / Cast(metarn['total'], FloatField()) * 100)
            rota = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalRota=Sum('rota2'), av_rota=Sum('rota2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
            apo = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalApo=Sum('apo3'), av_apo=Sum('apo3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
            penta = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalPenta=Sum('penta3'), av_penta=Sum('penta3', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
            infl2 = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalInfl2=Sum('infl2'), av_infl2=Sum('infl2', output_field=FloatField()) / Cast(metaLessyear['total'], FloatField()) * 100)
            neumo3 = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalNeumo3=Sum('neumo3'), av_neumo3=Sum('neumo3', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
            var = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalVar=Sum('varicela1'), av_var=Sum('varicela1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
            spr1 = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalSpr1=Sum('spr1'), av_spr1=Sum('spr1', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
            ama = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalAma=Sum('ama'), av_ama=Sum('ama', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
            hav = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalHav=Sum('hav'), av_hav=Sum('hav', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
            spr2 = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalSpr2=Sum('spr2'), av_spr2=Sum('spr2', output_field=FloatField()) / Cast(metaOneyear['total'], FloatField()) * 100)
            dpt2_ref = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalDpt2=Sum('dpt2_ref'), av_dpt2=Sum('dpt2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
            apo2_ref = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalApo2=Sum('apo2_ref'), av_apo2=Sum('apo2_ref', output_field=FloatField()) / Cast(metaFouryear['total'], FloatField()) * 100)
            vph_girl = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalGirl=Sum('vph_girl'), av_girl=Sum('vph_girl', output_field=FloatField()) / Cast(metaVphGirl['total'], FloatField()) * 100)
            vph_boy = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalBoy=Sum('vph_boy'), av_boy=Sum('vph_boy', output_field=FloatField()) / Cast(metaVphBoy['total'], FloatField()) * 100)
            gestante = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalDpta=Sum('dpta'), av_gest=Sum('dpta', output_field=FloatField()) / Cast(metaGest['total'], FloatField()) * 100)
            infAdult = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalInflAdult=Sum('infl_adult'), av_infAdul=Sum('infl_adult', output_field=FloatField()) / Cast(metaInfAdult['total'], FloatField()) * 100)
            neumoAdult = Coverage.objects.filter(anio=today.year, mes__range=[1, request.GET['month']], cod_eess=request.GET['eess']).aggregate(totalNeumoAdult=Sum('neumo_adult'), av_NeumoAdul=Sum('neumo_adult', output_field=FloatField()) / Cast(metaNeumoAdult['total'], FloatField()) * 100)

            resulTotal = {
                'av_bcg': bcg['av_bcg'], 'av_hvb': hvb['av_hvb'], 'av_rota': rota['av_rota'],
                'av_apo': apo['av_apo'], 'av_penta': penta['av_penta'], 'av_infl2': infl2['av_infl2'],
                'av_neumo3': neumo3['av_neumo3'], 'av_var': var['av_var'], 'av_spr1': spr1['av_spr1'],
                'av_ama': ama['av_ama'], 'av_hav': hav['av_hav'], 'av_spr2': spr2['av_spr2'], 'av_dpt2': dpt2_ref['av_dpt2'],
                'av_apo2': apo2_ref['av_apo2'], 'av_girl': vph_girl['av_girl'], 'av_boy': vph_boy['av_boy'],
                'av_gest': gestante['av_gest'], 'av_infAdul': infAdult['av_infAdul'], 'av_NeumoAdul': neumoAdult['av_NeumoAdul'],
            }

        dataList.extend([resulTotal])

        return HttpResponse(json.dumps(dataList), content_type='application/json')


class PrintNominal(TemplateView):
    def get(self, request, *args, **kwargs):
        locale.setlocale(locale.LC_TIME, 'es_ES')
        nameMonth = datetime.date(1900, int(request.GET['mes']), 1).strftime('%B')

        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:AW2", "medium", "57267C")
        set_border(self, ws, "A4:AW4", "medium", "366092")
        set_border(self, ws, "A6:AW6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws.merge_cells('B2:AW2')
        ws.row_dimensions[2].height = 23

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30

        ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'ESSALUD PASCO: COBERTURAS DE VACUNACIÓN - ENERO A ' + str(nameMonth).upper() + ' ' + str((date.today()).year)

        ws.merge_cells('A4:AW4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['A4'] = 'CODIFICACION: HVB: 90744   -   BCG: 90585'

        ws.merge_cells('A6:AW6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: ESSALUD con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws.merge_cells('A8:A9')
        ws['A8'] = '#'
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        ws.merge_cells('B8:B9')
        ws['B8'] = 'Centro Asistencial'
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B8'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        ws.merge_cells('C8:C9')
        ws['C8'] = 'Meta'
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws.merge_cells('D8:G8')
        ws['D8'] = 'Recien Nacidos'
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')
        ws['D9'] = 'BCG'
        ws['D9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D9'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')
        ws['E9'] = '%'
        ws['E9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E9'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')
        ws['F9'] = 'HVB'
        ws['F9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F9'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')
        ws['G9'] = '%'
        ws['G9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G9'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws.merge_cells('H8:H9')
        ws['H8'] = 'Meta'
        ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H8'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')

        ws.merge_cells('I8:P8')
        ws['I8'] = 'Menores de 1 Año'
        ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I8'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['I9'] = 'Rotavirus'
        ws['I9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['J9'] = '%'
        ws['J9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['K9'] = 'Polio'
        ws['K9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['L9'] = '%'
        ws['L9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['M9'] = 'Penta'
        ws['M9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['N9'] = '%'
        ws['N9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['O9'] = 'Influenza'
        ws['O9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')
        ws['P9'] = '%'
        ws['P9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P9'].fill = PatternFill(start_color='F9D3B9', end_color='F9D3B9', fill_type='solid')

        ws.merge_cells('Q8:Q9')
        ws['Q8'] = 'Meta'
        ws['Q8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q8'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')

        ws.merge_cells('R8:AC8')
        ws['R8'] = 'Niños de 1 Año'
        ws['R8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R8'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['R9'] = 'Neumo'
        ws['R9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['S9'] = '%'
        ws['S9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['T9'] = 'Varicela'
        ws['T9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['U9'] = '%'
        ws['U9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['V9'] = 'SPR1'
        ws['V9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['W9'] = '%'
        ws['W9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['X9'] = 'AMA'
        ws['X9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['X9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['Y9'] = '%'
        ws['Y9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Y9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Y9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['Z9'] = 'Hepatitis'
        ws['Z9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Z9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Z9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['AA9'] = '%'
        ws['AA9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AA9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AA9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['AB9'] = 'SPR2 '
        ws['AB9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AB9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AB9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')
        ws['AC9'] = '%'
        ws['AC9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AC9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AC9'].fill = PatternFill(start_color='F9D5FB', end_color='F9D5FB', fill_type='solid')

        ws.merge_cells('AD8:AD9')
        ws['AD8'] = 'Meta'
        ws['AD8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AD8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AD8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws.merge_cells('AE8:AH8')
        ws['AE8'] = 'Niños de 4 Años'
        ws['AE8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AE8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AE8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ws['AE9'] = 'DPT2'
        ws['AE9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AE9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AE9'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ws['AF9'] = '%'
        ws['AF9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AF9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AF9'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ws['AG9'] = 'APO2'
        ws['AG9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AG9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AG9'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ws['AH9'] = '%'
        ws['AH9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AH9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AH9'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws.merge_cells('AI8:AN8')
        ws['AI8'] = 'Niños y Niñas de 9 a 13 Años'
        ws['AI8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AI8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AI8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')
        ws['AI9'] = 'Meta'
        ws['AI9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AI9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AI9'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')
        ws['AJ9'] = 'VPH Niñas'
        ws['AJ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AJ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AJ9'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')
        ws['AK9'] = '%'
        ws['AK9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AK9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AK9'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')
        ws['AL9'] = 'Meta'
        ws['AL9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AL9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AL9'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')
        ws['AM9'] = 'VPH Niños'
        ws['AM9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AM9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AM9'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')
        ws['AN9'] = '%'
        ws['AN9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AN9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AN9'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws.merge_cells('AO8:AO9')
        ws['AO8'] = 'Meta'
        ws['AO8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AO8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AO8'].fill = PatternFill(start_color='B3F5C2', end_color='B3F5C2', fill_type='solid')

        ws.merge_cells('AP8:AQ8')
        ws['AP8'] = 'Gestantes'
        ws['AP8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AP8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AP8'].fill = PatternFill(start_color='B3F5C2', end_color='B3F5C2', fill_type='solid')
        ws['AP9'] = 'DPTA'
        ws['AP9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AP9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AP9'].fill = PatternFill(start_color='B3F5C2', end_color='B3F5C2', fill_type='solid')
        ws['AQ9'] = '%'
        ws['AQ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AQ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AQ9'].fill = PatternFill(start_color='B3F5C2', end_color='B3F5C2', fill_type='solid')

        ws.merge_cells('AR8:AW8')
        ws['AR8'] = 'Adulto Mayor'
        ws['AR8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AR8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AR8'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['AR9'] = 'Meta'
        ws['AR9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AR9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AR9'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['AS9'] = 'Influenza'
        ws['AS9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AS9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AS9'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['AT9'] = '%'
        ws['AT9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AT9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AT9'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['AU9'] = 'Meta'
        ws['AU9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AU9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AU9'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['AV9'] = 'Neumococo'
        ws['AV9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AV9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AV9'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['AW9'] = '%'
        ws['AW9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AW9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AW9'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        c = connection.cursor()
        if request.GET['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                c.execute("""SELECT cod_prov, provincia, cod_dist, distrito, cod_eess, eess, SUM(bcg) bcg, SUM(hvb) hvb, SUM(rota2) rota2, SUM(apo3) apo3, SUM(penta3) penta3,
                        SUM(infl2) infl2, sum(neumo3) neumo3, sum(varicela1) varicel, sum(spr1) spr1, sum(ama) ama, sum(hav) hav, sum(spr2) spr2,
                        sum(dpt2_ref) dpt2_ref, sum(apo2_ref) apo2_ref, sum(vph_girl) vph_girl, sum(vph_boy) vph_boy, sum(dpta) dpta,
                        sum(infl_adult) infl_adult, sum(neumo_adult) neumo_adult
                        INTO ESSALUD.dbo.printCobertNominal
                        FROM ESSALUD.dbo.boards_coverage WHERE Anio=%s and (Mes BETWEEN 1 and %s) and cod_eess=%s
                        AND provincia IS NOT NULL
                        GROUP BY cod_prov, provincia, cod_dist, distrito, cod_eess, eess
                        ORDER BY cod_prov, provincia, cod_dist, distrito""" % ((date.today()).year, request.GET['mes'], self.request.session['sytem']['codeca']))

            elif self.request.session['sytem']['typeca'] == 'DS':
                c.execute("""SELECT cod_prov, provincia, cod_dist, distrito, cod_eess, eess, SUM(bcg) bcg, SUM(hvb) hvb, SUM(rota2) rota2, SUM(apo3) apo3, SUM(penta3) penta3,
                        SUM(infl2) infl2, sum(neumo3) neumo3, sum(varicela1) varicel, sum(spr1) spr1, sum(ama) ama, sum(hav) hav, sum(spr2) spr2,
                        sum(dpt2_ref) dpt2_ref, sum(apo2_ref) apo2_ref, sum(vph_girl) vph_girl, sum(vph_boy) vph_boy, sum(dpta) dpta,
                        sum(infl_adult) infl_adult, sum(neumo_adult) neumo_adult
                        INTO ESSALUD.dbo.printCobertNominal
                        FROM ESSALUD.dbo.boards_coverage WHERE Anio=%s and (Mes BETWEEN 1 and %s) and cod_dist=%s
                        AND provincia IS NOT NULL
                        GROUP BY cod_prov, provincia, cod_dist, distrito, cod_eess, eess
                        ORDER BY cod_prov, provincia, cod_dist, distrito""" % ((date.today()).year, request.GET['mes'], self.request.session['sytem']['codeca']))

            elif self.request.session['sytem']['typeca'] == 'PR':
                c.execute("""SELECT cod_prov, provincia, cod_dist, distrito, cod_eess, eess, SUM(bcg) bcg, SUM(hvb) hvb, SUM(rota2) rota2, SUM(apo3) apo3, SUM(penta3) penta3,
                        SUM(infl2) infl2, sum(neumo3) neumo3, sum(varicela1) varicel, sum(spr1) spr1, sum(ama) ama, sum(hav) hav, sum(spr2) spr2,
                        sum(dpt2_ref) dpt2_ref, sum(apo2_ref) apo2_ref, sum(vph_girl) vph_girl, sum(vph_boy) vph_boy, sum(dpta) dpta,
                        sum(infl_adult) infl_adult, sum(neumo_adult) neumo_adult
                        INTO ESSALUD.dbo.printCobertNominal
                        FROM ESSALUD.dbo.boards_coverage WHERE Anio=%s and (Mes BETWEEN 1 and %s) and cod_prov=%s
                        AND provincia IS NOT NULL
                        GROUP BY cod_prov, provincia, cod_dist, distrito, cod_eess, eess
                        ORDER BY cod_prov, provincia, cod_dist, distrito""" % ((date.today()).year, request.GET['mes'], self.request.session['sytem']['codeca']))

            elif self.request.session['sytem']['typeca'] == 'DP':
                c.execute("""SELECT cod_prov, provincia, cod_dist, distrito, cod_eess, eess, SUM(bcg) bcg, SUM(hvb) hvb, SUM(rota2) rota2, SUM(apo3) apo3, SUM(penta3) penta3,
                        SUM(infl2) infl2, sum(neumo3) neumo3, sum(varicela1) varicel, sum(spr1) spr1, sum(ama) ama, sum(hav) hav, sum(spr2) spr2,
                        sum(dpt2_ref) dpt2_ref, sum(apo2_ref) apo2_ref, sum(vph_girl) vph_girl, sum(vph_boy) vph_boy, sum(dpta) dpta,
                        sum(infl_adult) infl_adult, sum(neumo_adult) neumo_adult
                        INTO ESSALUD.dbo.printCobertNominal
                        FROM ESSALUD.dbo.boards_coverage WHERE Anio=%s and (Mes BETWEEN 1 and %s) and cod_dep=%s
                        AND provincia IS NOT NULL
                        GROUP BY cod_prov, provincia, cod_dist, distrito, cod_eess, eess
                        ORDER BY cod_prov, provincia, cod_dist, distrito""" % ((date.today()).year, request.GET['mes'], self.request.session['sytem']['codeca']))

        else:
            c.execute("""SELECT cod_prov, provincia, cod_dist, distrito, cod_eess, eess, SUM(bcg) bcg, SUM(hvb) hvb, SUM(rota2) rota2, SUM(apo3) apo3, SUM(penta3) penta3,
                        SUM(infl2) infl2, sum(neumo3) neumo3, sum(varicela1) varicel, sum(spr1) spr1, sum(ama) ama, sum(hav) hav, sum(spr2) spr2,
                        sum(dpt2_ref) dpt2_ref, sum(apo2_ref) apo2_ref, sum(vph_girl) vph_girl, sum(vph_boy) vph_boy, sum(dpta) dpta,
                        sum(infl_adult) infl_adult, sum(neumo_adult) neumo_adult
                        INTO ESSALUD.dbo.printCobertNominal
                        FROM ESSALUD.dbo.boards_coverage WHERE Anio=%s and (Mes BETWEEN 1 and %s) and cod_eess=%s
                        AND provincia IS NOT NULL
                        GROUP BY cod_prov, provincia, cod_dist, distrito, cod_eess, eess
                        ORDER BY cod_prov, provincia, cod_dist, distrito""" % ((date.today()).year, request.GET['mes'], request.GET['eess']))

        c.execute("""SELECT B.provincia, B.distrito, B.establecimiento, SUM(B.rn) meta_rn, iif(A.bcg is null, 0, a.bcg) tot_bcg, iif(A.hvb is null, 0, a.hvb) tot_hvb,
                    round(cast(iif(A.bcg is null, 0, a.bcg) as float)/cast(SUM(B.rn) as float) * 100,1) av_bcg,
                        round(cast(iif(A.hvb is null, 0, a.hvb) as float)/cast(SUM(B.rn) as float) * 100,1) av_hvb,
                        SUM(B.less_1year) meta_lessyear, iif(A.rota2 is null, 0, a.rota2) tot_rota2, iif(A.apo3 is null, 0, a.apo3) tot_apo3,
                        iif(A.penta3 is null, 0, a.penta3) tot_penta3, iif(A.infl2 is null, 0, a.infl2) tot_infl2,
                        round(cast(iif(A.rota2 is null, 0, a.rota2) as float)/cast(SUM(B.less_1year) as float) * 100,1) av_rota2,
                        round(cast(iif(A.apo3 is null, 0, a.apo3) as float)/cast(SUM(B.less_1year) as float) * 100,1) av_apo3,
                        round(cast(iif(A.penta3 is null, 0, a.penta3) as float)/cast(SUM(B.less_1year) as float) * 100,1) av_penta3,
                        round(cast(iif(A.infl2 is null, 0, a.infl2) as float)/cast(SUM(B.less_1year) as float) * 100,1) av_infl2,
                        SUM(B.one_year) meta_oneyear, iif(A.neumo3 is null, 0, a.neumo3) tot_neumo3, iif(A.varicel is null, 0, a.varicel) tot_varicel,
                        iif(A.spr1 is null, 0, a.spr1) tot_spr1, iif(A.ama is null, 0, a.ama) tot_ama, iif(A.hav is null, 0, a.hav) tot_hav,
                        iif(A.spr2 is null, 0, a.spr2) tot_spr2, round(cast(iif(A.neumo3 is null, 0, a.neumo3) as float)/cast(SUM(B.one_year) as float) * 100,1) av_neumo3,
                        round(cast(iif(A.varicel is null, 0, a.varicel) as float)/cast(SUM(B.one_year) as float) * 100,1) av_varicel,
                        round(cast(iif(A.spr1 is null, 0, a.spr1) as float)/cast(SUM(B.one_year) as float) * 100,1) av_spr1,
                        round(cast(iif(A.ama is null, 0, a.ama) as float)/cast(SUM(B.one_year) as float) * 100,1) av_ama,
                        round(cast(iif(A.hav is null, 0, a.hav) as float)/cast(SUM(B.one_year) as float) * 100,1) av_hav,
                        round(cast(iif(A.spr2 is null, 0, a.spr2) as float)/cast(SUM(B.one_year) as float) * 100,1) av_infl2spr2,
                        SUM(B.four_year) meta_fouryear, iif(A.dpt2_ref is null, 0, a.dpt2_ref) tot_dpt2_ref, iif(A.apo2_ref is null, 0, a.apo2_ref) tot_apo2_ref,
                        round(cast(iif(A.dpt2_ref is null, 0, a.dpt2_ref) as float)/cast(SUM(B.four_year) as float) * 100,1) av_dpt2_ref,
                        round(cast(iif(A.apo2_ref is null, 0, a.apo2_ref) as float)/cast(SUM(B.four_year) as float) * 100,1) av_apo2_ref,
                        SUM(B.girl9_13) meta_girl, iif(A.vph_girl is null, 0, a.vph_girl) tot_vph_girl, SUM(B.boy9_13) meta_boy,
                        iif(A.vph_boy is null, 0, a.vph_boy) tot_vph_boy,
                        round(cast(iif(A.vph_girl is null, 0, a.vph_girl) as float)/cast(SUM(B.girl9_13) as float) * 100,1) av_vph_girl,
                        round(cast(iif(A.vph_boy is null, 0, a.vph_boy) as float)/cast(SUM(B.boy9_13) as float) * 100,1) av_vph_boy,
                        SUM(B.pregnant) meta_girl, iif(A.dpta is null, 0, a.dpta) tot_dpta,
                        round(cast(iif(A.dpta is null, 0, a.dpta) as float)/cast(SUM(B.pregnant) as float) * 100,1) av_dpta,
                        SUM(B.adult60) meta_infl, iif(A.infl_adult is null, 0, a.infl_adult) tot_infl_adult, SUM(B.adult30) meta_neumo,
                        iif(A.neumo_adult is null, 0, a.neumo_adult) tot_neumo_adult,
                        round(cast(iif(A.infl_adult is null, 0, a.infl_adult) as float)/cast(SUM(B.adult60) as float) * 100,1) av_infl_adult,
                        round(cast(iif(A.neumo_adult is null, 0, a.neumo_adult) as float)/cast(SUM(B.adult30) as float) * 100,1) av_neumo_adult
                    FROM ESSALUD.dbo.boards_metacoverage AS B RIGHT JOIN ESSALUD.dbo.printCobertNominal A
                    ON A.cod_eess=B.cod_eess where b.anio=%s
                    GROUP BY B.provincia, B.distrito, B.establecimiento, A.bcg, A.hvb, a.rota2, a.apo3, a.penta3, a.infl2, a.neumo3, a.varicel, a.spr1, a.ama, a.hav, a.spr2,
                    dpt2_ref, apo2_ref, a.vph_girl, a.vph_boy, a.dpta, a.infl_adult, a.neumo_adult
                    DROP TABLE ESSALUD.dbo.printCobertNominal""" % ((date.today()).year))

        cont = 10
        num = 1
        data = []
        for a in c.fetchall():
            datos = {'provincia': a[0], 'distrito': a[1], 'establecimiento': a[2], 'meta_rn': a[3], 'bcg': 0 if a[4] == None else a[4], 'hvb': 0 if a[5] == None else a[5],
                     'av_bcg': a[6], 'av_hvb': a[7], 'meta_lessyear': a[8], 'rota2': 0 if a[9] == None else a[9], 'apo3': 0 if a[10] == None else a[10],
                     'penta3': 0 if a[11] == None else a[11], 'infl2': 0 if a[12] == None else a[12], 'av_rota2': a[13], 'av_apo3': a[14], 'av_penta3': a[15],
                     'av_infl2': a[16], 'meta_oneyear': a[17], 'neumo3': 0 if a[18] == None else a[18], 'varicel': 0 if a[19] == None else a[19],
                     'spr1': 0 if a[20] == None else a[20], 'ama': 0 if a[21] == None else a[21], 'hav': 0 if a[22] == None else a[22], 'spr2': 0 if a[23] == None else a[23],
                     'av_neumo3': a[24], 'av_varicel': a[25], 'av_spr1': a[26], 'av_ama': a[27], 'av_hav': a[28], 'av_spr2': a[29], 'meta_fouryear': a[30],
                     'dpt2': 0 if a[31] == None else a[31], 'apo2': 0 if a[32] == None else a[32], 'av_dpt2': a[33], 'av_apo2': a[34], 'meta_girl': a[35],
                     'girl': 0 if a[36] == None else a[36], 'meta_boy': a[37], 'boy': 0 if a[38] == None else a[38], 'av_girl': a[39], 'av_boy': a[40],
                     'meta_dpta': a[41], 'dpta': 0 if a[42] == None else a[42], 'av_dpta': a[43], 'meta_infl': a[44], 'infl_adult': 0 if a[45] == None else a[45],
                     'meta_neumo': a[46], 'neumo_adult': 0 if a[47] == None else a[47], 'av_infl': a[48], 'av_adul': a[49],
                    }

            data.append(datos)

        for nom in data:
            ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=1).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=1).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=1).value = num

            ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=2).value = nom['establecimiento']

            ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=3).value = nom['meta_rn']

            ws.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=4).value = nom['bcg']

            ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            traffic = '⬤'
            if nom['av_bcg'] < 41:
                ws.cell(row=cont, column=5).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_bcg'] > 40 and nom['av_bcg'] < 61:
                ws.cell(row=cont, column=5).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_bcg'] > 60 and nom['av_bcg'] < 95:
                ws.cell(row=cont, column=5).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_bcg'] > 94:
                ws.cell(row=cont, column=5).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=5).value = str(nom['av_bcg'])+' % '+traffic

            ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=6).value = nom['hvb']

            ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=9)
            if nom['av_hvb'] < 41:
                ws.cell(row=cont, column=7).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_hvb'] > 40 and nom['av_hvb'] < 61:
                ws.cell(row=cont, column=7).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_hvb'] > 60 and nom['av_hvb'] < 95:
                ws.cell(row=cont, column=7).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_hvb'] > 94:
                ws.cell(row=cont, column=7).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=7).value = str(nom['av_hvb'])+' % '+traffic

            ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=8).value = nom['meta_lessyear']

            ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=9).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=9).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=9).value = nom['rota2']

            ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=10).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=10).font = Font(name='Calibri', size=9)
            if nom['av_rota2'] < 41:
                ws.cell(row=cont, column=10).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_rota2'] > 40 and nom['av_rota2'] < 61:
                ws.cell(row=cont, column=10).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_rota2'] > 60 and nom['av_rota2'] < 95:
                ws.cell(row=cont, column=10).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_rota2'] > 94:
                ws.cell(row=cont, column=10).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=10).value = str(nom['av_rota2'])+' % '+traffic

            ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=11).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=11).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=11).value = nom['apo3']

            ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=12).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=12).font = Font(name='Calibri', size=9)
            if nom['av_apo3'] < 41:
                ws.cell(row=cont, column=12).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_apo3'] > 40 and nom['av_apo3'] < 61:
                ws.cell(row=cont, column=12).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_apo3'] > 60 and nom['av_apo3'] < 95:
                ws.cell(row=cont, column=12).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_apo3'] > 94:
                ws.cell(row=cont, column=12).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=12).value = str(nom['av_apo3'])+' % '+traffic

            ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=13).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=13).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=13).value = nom['penta3']

            ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=14).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=14).font = Font(name='Calibri', size=9)
            if nom['av_penta3'] < 41:
                ws.cell(row=cont, column=14).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_penta3'] > 40 and nom['av_penta3'] < 61:
                ws.cell(row=cont, column=14).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_penta3'] > 60 and nom['av_penta3'] < 95:
                ws.cell(row=cont, column=14).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_penta3'] > 94:
                ws.cell(row=cont, column=14).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=14).value = str(nom['av_penta3'])+' % '+traffic

            ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=15).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=15).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=15).value = nom['infl2']

            ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=16).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=16).font = Font(name='Calibri', size=9)
            if nom['av_infl2'] < 41:
                ws.cell(row=cont, column=16).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_infl2'] > 40 and nom['av_infl2'] < 61:
                ws.cell(row=cont, column=16).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_infl2'] > 60 and nom['av_infl2'] < 95:
                ws.cell(row=cont, column=16).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_infl2'] > 94:
                ws.cell(row=cont, column=16).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=16).value = str(nom['av_infl2'])+' % '+traffic

            ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=17).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=17).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=17).value = nom['meta_oneyear']

            ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=18).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=18).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=18).value = nom['neumo3']

            ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=19).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=19).font = Font(name='Calibri', size=9)
            if nom['av_neumo3'] < 41:
                ws.cell(row=cont, column=19).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_neumo3'] > 40 and nom['av_neumo3'] < 61:
                ws.cell(row=cont, column=19).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_neumo3'] > 60 and nom['av_neumo3'] < 95:
                ws.cell(row=cont, column=19).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_neumo3'] > 94:
                ws.cell(row=cont, column=19).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=19).value = str(nom['av_neumo3'])+' % '+traffic

            ws.cell(row=cont, column=20).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=20).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=20).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=20).value = nom['varicel']

            ws.cell(row=cont, column=21).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=21).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=21).font = Font(name='Calibri', size=9)
            if nom['av_varicel'] < 41:
                ws.cell(row=cont, column=21).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_varicel'] > 40 and nom['av_varicel'] < 61:
                ws.cell(row=cont, column=21).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_varicel'] > 60 and nom['av_varicel'] < 95:
                ws.cell(row=cont, column=21).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_varicel'] > 94:
                ws.cell(row=cont, column=21).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=21).value = str(nom['av_varicel'])+' % '+traffic

            ws.cell(row=cont, column=22).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=22).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=22).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=22).value = nom['spr1']

            ws.cell(row=cont, column=23).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=23).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=23).font = Font(name='Calibri', size=9)
            if nom['av_spr1'] < 41:
                ws.cell(row=cont, column=23).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_spr1'] > 40 and nom['av_spr1'] < 61:
                ws.cell(row=cont, column=23).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_spr1'] > 60 and nom['av_spr1'] < 95:
                ws.cell(row=cont, column=23).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_spr1'] > 94:
                ws.cell(row=cont, column=23).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=23).value = str(nom['av_spr1'])+' % '+traffic

            ws.cell(row=cont, column=24).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=24).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=24).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=24).value = nom['ama']

            ws.cell(row=cont, column=25).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=25).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=25).font = Font(name='Calibri', size=9)
            if nom['av_ama'] < 41:
                ws.cell(row=cont, column=25).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_ama'] > 40 and nom['av_ama'] < 61:
                ws.cell(row=cont, column=25).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_ama'] > 60 and nom['av_ama'] < 95:
                ws.cell(row=cont, column=25).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_ama'] > 94:
                ws.cell(row=cont, column=25).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=25).value = str(nom['av_ama'])+' % '+traffic

            ws.cell(row=cont, column=26).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=26).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=26).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=26).value = nom['hav']

            ws.cell(row=cont, column=27).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=27).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=27).font = Font(name='Calibri', size=9)
            if nom['av_hav'] < 41:
                ws.cell(row=cont, column=27).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_hav'] > 40 and nom['av_hav'] < 61:
                ws.cell(row=cont, column=27).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_hav'] > 60 and nom['av_hav'] < 95:
                ws.cell(row=cont, column=27).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_hav'] > 94:
                ws.cell(row=cont, column=27).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=27).value = str(nom['av_hav'])+' % '+traffic

            ws.cell(row=cont, column=28).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=28).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=28).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=28).value = nom['spr2']

            ws.cell(row=cont, column=29).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=29).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=29).font = Font(name='Calibri', size=9)
            if nom['av_spr2'] < 41:
                ws.cell(row=cont, column=29).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_spr2'] > 40 and nom['av_spr2'] < 61:
                ws.cell(row=cont, column=29).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_spr2'] > 60 and nom['av_spr2'] < 95:
                ws.cell(row=cont, column=29).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_spr2'] > 94:
                ws.cell(row=cont, column=29).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=29).value = str(nom['av_spr2'])+' % '+traffic

            ws.cell(row=cont, column=30).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=30).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=30).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=30).value = nom['meta_fouryear']

            ws.cell(row=cont, column=31).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=31).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=31).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=31).value = nom['dpt2']

            ws.cell(row=cont, column=32).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=32).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=32).font = Font(name='Calibri', size=9)
            if nom['av_dpt2'] < 41:
                ws.cell(row=cont, column=32).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_dpt2'] > 40 and nom['av_dpt2'] < 61:
                ws.cell(row=cont, column=32).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_dpt2'] > 60 and nom['av_dpt2'] < 95:
                ws.cell(row=cont, column=32).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_dpt2'] > 94:
                ws.cell(row=cont, column=32).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=32).value = str(nom['av_dpt2'])+' % '+traffic

            ws.cell(row=cont, column=33).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=33).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=33).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=33).value = nom['apo2']

            ws.cell(row=cont, column=34).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=34).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=34).font = Font(name='Calibri', size=9)
            if nom['av_apo2'] < 41:
                ws.cell(row=cont, column=34).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_apo2'] > 40 and nom['av_apo2'] < 61:
                ws.cell(row=cont, column=34).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_apo2'] > 60 and nom['av_apo2'] < 95:
                ws.cell(row=cont, column=34).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_apo2'] > 94:
                ws.cell(row=cont, column=34).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=34).value = str(nom['av_apo2'])+' % '+traffic

            ws.cell(row=cont, column=35).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=35).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=35).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=35).value = nom['meta_girl']

            ws.cell(row=cont, column=36).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=36).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=36).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=36).value = nom['girl']

            ws.cell(row=cont, column=37).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=37).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=37).font = Font(name='Calibri', size=9)
            if nom['av_girl'] < 41:
                ws.cell(row=cont, column=37).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_girl'] > 40 and nom['av_girl'] < 61:
                ws.cell(row=cont, column=37).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_girl'] > 60 and nom['av_girl'] < 95:
                ws.cell(row=cont, column=37).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_girl'] > 94:
                ws.cell(row=cont, column=37).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=37).value = str(nom['av_girl'])+' % '+traffic

            ws.cell(row=cont, column=38).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=38).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=38).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=38).value = nom['meta_boy']

            ws.cell(row=cont, column=39).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=39).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=39).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=39).value = nom['boy']

            ws.cell(row=cont, column=40).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=40).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=40).font = Font(name='Calibri', size=9)
            if nom['av_boy'] < 41:
                ws.cell(row=cont, column=40).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_boy'] > 40 and nom['av_boy'] < 61:
                ws.cell(row=cont, column=40).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_boy'] > 60 and nom['av_boy'] < 95:
                ws.cell(row=cont, column=40).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_boy'] > 94:
                ws.cell(row=cont, column=40).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=40).value = str(nom['av_boy'])+' % '+traffic

            ws.cell(row=cont, column=41).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=41).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=41).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=41).value = nom['meta_dpta']

            ws.cell(row=cont, column=42).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=42).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=42).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=42).value = nom['dpta']

            ws.cell(row=cont, column=43).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=43).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=43).font = Font(name='Calibri', size=9)
            if nom['av_dpta'] < 41:
                ws.cell(row=cont, column=43).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_dpta'] > 40 and nom['av_dpta'] < 61:
                ws.cell(row=cont, column=43).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_dpta'] > 60 and nom['av_dpta'] < 95:
                ws.cell(row=cont, column=43).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_dpta'] > 94:
                ws.cell(row=cont, column=43).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=43).value = str(nom['av_dpta'])+' % '+traffic

            ws.cell(row=cont, column=44).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=44).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=44).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=44).value = nom['meta_infl']

            ws.cell(row=cont, column=45).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=45).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=45).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=45).value = nom['infl_adult']

            ws.cell(row=cont, column=46).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=46).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=46).font = Font(name='Calibri', size=9)
            if nom['av_infl'] < 41:
                ws.cell(row=cont, column=46).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_infl'] > 40 and nom['av_infl'] < 61:
                ws.cell(row=cont, column=46).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_infl'] > 60 and nom['av_infl'] < 95:
                ws.cell(row=cont, column=46).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_infl'] > 94:
                ws.cell(row=cont, column=46).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=46).value = str(nom['av_infl'])+' % '+traffic

            ws.cell(row=cont, column=47).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=47).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=47).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=47).value = nom['meta_neumo']

            ws.cell(row=cont, column=48).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=48).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=48).font = Font(name='Calibri', size=9)
            ws.cell(row=cont, column=48).value = nom['neumo_adult']

            ws.cell(row=cont, column=49).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=49).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws.cell(row=cont, column=49).font = Font(name='Calibri', size=9)
            if nom['av_adul'] < 41:
                ws.cell(row=cont, column=49).font = Font(name='Calibri', size=10, color='FF2929')
            elif nom['av_adul'] > 40 and nom['av_adul'] < 61:
                ws.cell(row=cont, column=49).font = Font(name='Calibri', size=10, color='FF9900')
            elif nom['av_adul'] > 60 and nom['av_adul'] < 95:
                ws.cell(row=cont, column=49).font = Font(name='Calibri', size=10, color='00B050')
            elif nom['av_adul'] > 94:
                ws.cell(row=cont, column=49).font = Font(name='Calibri', size=10, color='4472C4')

            ws.cell(row=cont, column=49).value = str(nom['av_adul'])+' % '+traffic

            cont = cont+1
            num = num+1

        nombre_archivo = "DEIT_PASCO COBERTURA DE VACUNACION 2024.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'COBERTURA DE VACUNACION'
        wb.save(response)
        return response

