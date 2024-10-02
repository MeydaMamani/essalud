from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView, View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, QueryDict
from django.db.models import Case, When, IntegerField, FloatField, ExpressionWrapper, Q, F, Sum, Count, IntegerField, Avg, Value, DecimalField
from django.db.models.functions import Cast, Round
import json
from datetime import date, datetime

from .models import PackChildFollow, PregnantFollow
from apps.main.models import Provincia, Distrito, Establecimiento

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color

import locale
import datetime

class FollowKidsView(TemplateView):
    template_name = 'boys/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.session['sytem']['typeca'] == 'CA':
            context['establecimiento'] = Establecimiento.objects.filter(codigo=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DS':
            context['establecimiento'] = Establecimiento.objects.filter(dist_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'PR':
            context['establecimiento'] = Establecimiento.objects.filter(prov_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DP':
            context['establecimiento'] = Establecimiento.objects.filter(dep_id=self.request.session['sytem']['codeca'])
        return context


class ListKidsFollow(View):
    http_method_names = ['get', 'post', 'put', 'delete']
    def post(self,request,*args, **kwargs):
        json_data4 = []
        if len(request.POST['mes']) == 1:
            mes = '0'+request.POST['mes']
        else:
            mes = request.POST['mes']

        if request.POST['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                total = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif self.request.session['sytem']['typeca'] == 'DS':
                total = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif self.request.session['sytem']['typeca'] == 'PR':
                total = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif self.request.session['sytem']['typeca'] == 'DP':
                total = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            total = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).aggregate(total=Sum('den'))['total']
            cumplen = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).aggregate(cumplen=Sum('num'))['cumplen']
            dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
            dataProv = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).values('provincia').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataDist = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).values('establecimiento').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataNom = PackChildFollow.objects.filter(fec_nac__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).order_by('cod_eess')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)
        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintPackChild(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:BA2", "medium", "57267C")
        set_border(self, ws, "A4:BA4", "medium", "366092")
        set_border(self, ws, "A6:BA6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws.merge_cells('B2:BA2')
        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10

        ws.column_dimensions['AQ'].width = 6
        ws.column_dimensions['AR'].width = 6
        ws.column_dimensions['AS'].width = 6
        ws.column_dimensions['AT'].width = 6
        ws.column_dimensions['AU'].width = 6
        ws.column_dimensions['AV'].width = 6
        ws.column_dimensions['AW'].width = 6
        ws.column_dimensions['AX'].width = 6
        ws.column_dimensions['AY'].width = 6
        ws.column_dimensions['AZ'].width = 8
        ws.column_dimensions['BA'].width = 8
        ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'ESSALUD: Niñas y niños menores de 12 meses de edad, procedentes de los quintiles 1 y 2 de pobreza departamental que reciben el paquete integrado de servicios -' + request.GET['anio'] + ' de ' + nameMonth.upper() + ' A LA FECHA'

        ws.merge_cells('A4:BA4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['A4'] = 'CODIFICACION: Cred Rn: 99381.01 - Cred Mes: 99381 - Vacuna Antipolio: 90712 - Vacuna Pentavalente: 90722 - Dx Anemia: D500, D508, D509, D649, D539 - Suplementación: 99199.17, 99199.19 - Prematuros: P073, P071, P0711, P00712 - Dosaje Hemoglobina: 85018, 85018.01'

        ws.merge_cells('A6:BA6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: EsSalud con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A8'] = '#'
        ws['A8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['B8'] = 'Centro Asistencial'
        ws['B8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['C8'] = 'Documento'
        ws['C8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['D8'] = 'Apellidos y Nombres'
        ws['D8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['E8'] = 'Fecha Nacido'
        ws['E8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['F8'] = '1° Ctrl'
        ws['F8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['G8'] = '2° Ctrl'
        ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['H8'] = '3° Ctrl'
        ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['I8'] = '4° Ctrl'
        ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['J8'] = 'Ctrls Rn'
        ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J8'].fill = PatternFill(start_color='B3F5C2', end_color='B3F5C2', fill_type='solid')

        ws['K8'] = 'Cred 1'
        ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['L8'] = 'Cred 2'
        ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['M8'] = 'Antineumo 2M'
        ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['M8'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        ws['N8'] = 'Rotavirus 2M'
        ws['N8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N8'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        ws['O8'] = 'Antipolio 2M'
        ws['O8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['O8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws['P8'] = 'Penta 2M'
        ws['P8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['Q8'] = 'Cred 3'
        ws['Q8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['R8'] = 'Cred 4'
        ws['R8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['S8'] = 'Suple 4'
        ws['S8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['T8'] = 'Antineumo 4M'
        ws['T8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['T8'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        ws['U8'] = 'Rotavirus 4M'
        ws['U8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['U8'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        ws['V8'] = 'Penta 4M'
        ws['V8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['V8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['W8'] = 'Antipolio 4M'
        ws['W8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['W8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['W8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws['X8'] = 'Cred 5'
        ws['X8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['X8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['X8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['Y8'] = 'Suple 5'
        ws['Y8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Y8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Y8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Y8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['Z8'] = 'Cred 6'
        ws['Z8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Z8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Z8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Z8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AA8'] = 'Tamizaje'
        ws['AA8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AA8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AA8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AA8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AB8'] = 'Dx Anemia'
        ws['AB8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AB8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AB8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AB8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AC8'] = 'Suple 6'
        ws['AC8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AC8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AC8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AC8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AD8'] = 'Antipolio 6M'
        ws['AD8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AD8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AD8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AD8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws['AE8'] = 'Penta 6M'
        ws['AE8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AE8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AE8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AE8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AF8'] = 'Cred 7'
        ws['AF8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AF8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AF8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AF8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AG8'] = 'Suple 7'
        ws['AG8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AG8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AG8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AG8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AH8'] = 'Cred 8'
        ws['AH8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AH8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AH8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AH8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AI8'] = 'Suple 8'
        ws['AI8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AI8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AI8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AI8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AJ8'] = 'Cred 9'
        ws['AJ8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AJ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AJ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AJ8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AK8'] = 'Suple 9'
        ws['AK8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AK8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AK8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AK8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AL8'] = 'Cred 10'
        ws['AL8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AL8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AL8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AL8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AM8'] = 'Suple 10'
        ws['AM8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AM8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AM8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AM8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AN8'] = 'Cred 11'
        ws['AN8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AN8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AN8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AN8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AO8'] = 'Suple 11'
        ws['AO8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AO8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AO8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AO8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AP8'] = 'Eval Oral'
        ws['AP8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AP8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AP8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AP8'].fill = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')

        ws['AQ8'] = 'C1-C2'
        ws['AQ8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AQ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AQ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AQ8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AR8'] = 'C2-C3'
        ws['AR8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AR8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AR8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AR8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AS8'] = 'C3-C4'
        ws['AS8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AS8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AS8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AS8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AT8'] = 'C4-C5'
        ws['AT8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AT8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AT8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AT8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AU8'] = 'C5-C6'
        ws['AU8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AU8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AU8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AU8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AV8'] = 'C6-C7'
        ws['AV8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AV8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AV8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AV8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AW8'] = 'C7-C8'
        ws['AW8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AW8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AW8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AW8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AX8'] = 'C8-C9'
        ws['AX8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AX8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AX8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AX8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AY8'] = 'C9-C10'
        ws['AY8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AY8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AY8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AY8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AZ8'] = 'C10-C11'
        ws['AZ8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AZ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AZ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AZ8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['BA8'] = 'Cumple'
        ws['BA8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BA8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BA8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BA8'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        if len(request.GET['mes']) == 1:
            mes = '0'+request.GET['mes']
        else:
            mes = request.GET['mes']

        if request.GET['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                dataNom = PackChildFollow.objects.filter(cod_eess=self.request.session['sytem']['codeca'], fec_nac__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            if self.request.session['sytem']['typeca'] == 'DS':
                dataNom = PackChildFollow.objects.filter(cod_dist=self.request.session['sytem']['codeca'], fec_nac__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            if self.request.session['sytem']['typeca'] == 'PR':
                dataNom = PackChildFollow.objects.filter(cod_prov=self.request.session['sytem']['codeca'], fec_nac__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            if self.request.session['sytem']['typeca'] == 'DP':
                dataNom = PackChildFollow.objects.filter(cod_dep=self.request.session['sytem']['codeca'], fec_nac__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
        else:
            dataNom = PackChildFollow.objects.filter(cod_eess=request.GET['eess'], fec_nac__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 9
        cant = len(dataNom)
        num=1
        if cant > 0:
            for paqNinio in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = paqNinio['fields']['establecimiento']
                ws.cell(row=cont, column=3).value = paqNinio['fields']['documento']
                ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=4).value = paqNinio['fields']['ape_nombres']
                ws.cell(row=cont, column=5).value = paqNinio['fields']['fec_nac']
                ws.cell(row=cont, column=6).value = paqNinio['fields']['ctrl1rn']
                ws.cell(row=cont, column=7).value = paqNinio['fields']['ctrl2rn']
                ws.cell(row=cont, column=8).value = paqNinio['fields']['ctrl3rn']
                ws.cell(row=cont, column=9).value = paqNinio['fields']['ctrl4rn']

                if paqNinio['fields']['num_rn'] == '1':
                    cumplen = '✔'
                    ws.cell(row=cont, column=10).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=10).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = cumplen
                ws.cell(row=cont, column=11).value = paqNinio['fields']['cred1']
                ws.cell(row=cont, column=12).value = paqNinio['fields']['cred2']
                ws.cell(row=cont, column=13).value = paqNinio['fields']['neumo2']
                ws.cell(row=cont, column=14).value = paqNinio['fields']['rota2']
                ws.cell(row=cont, column=15).value = paqNinio['fields']['polio2']
                ws.cell(row=cont, column=16).value = paqNinio['fields']['penta2']
                ws.cell(row=cont, column=17).value = paqNinio['fields']['cred3']
                ws.cell(row=cont, column=18).value = paqNinio['fields']['cred4']
                ws.cell(row=cont, column=19).value = paqNinio['fields']['suple4']
                ws.cell(row=cont, column=20).value = paqNinio['fields']['neumo4']
                ws.cell(row=cont, column=21).value = paqNinio['fields']['rota4']
                ws.cell(row=cont, column=22).value = paqNinio['fields']['penta4']
                ws.cell(row=cont, column=23).value = paqNinio['fields']['polio4']
                ws.cell(row=cont, column=24).value = paqNinio['fields']['cred5']
                ws.cell(row=cont, column=25).value = paqNinio['fields']['suple5']
                ws.cell(row=cont, column=26).value = paqNinio['fields']['cred6']
                ws.cell(row=cont, column=27).value = paqNinio['fields']['tmz']
                ws.cell(row=cont, column=28).value = paqNinio['fields']['dxAnemia']
                ws.cell(row=cont, column=29).value = paqNinio['fields']['suple6']
                ws.cell(row=cont, column=30).value = paqNinio['fields']['polio6']
                ws.cell(row=cont, column=31).value = paqNinio['fields']['penta6']
                ws.cell(row=cont, column=32).value = paqNinio['fields']['cred7']
                ws.cell(row=cont, column=33).value = paqNinio['fields']['suple7']
                ws.cell(row=cont, column=34).value = paqNinio['fields']['cred8']
                ws.cell(row=cont, column=35).value = paqNinio['fields']['suple8']
                ws.cell(row=cont, column=36).value = paqNinio['fields']['cred9']
                ws.cell(row=cont, column=37).value = paqNinio['fields']['suple9']
                ws.cell(row=cont, column=38).value = paqNinio['fields']['cred10']
                ws.cell(row=cont, column=39).value = paqNinio['fields']['suple10']
                ws.cell(row=cont, column=40).value = paqNinio['fields']['cred11']
                ws.cell(row=cont, column=41).value = paqNinio['fields']['suple11']
                ws.cell(row=cont, column=42).value = paqNinio['fields']['eval_oral']
                ws.cell(row=cont, column=43).value = paqNinio['fields']['dif1']
                ws.cell(row=cont, column=43).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=44).value = paqNinio['fields']['dif2']
                ws.cell(row=cont, column=44).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=45).value = paqNinio['fields']['dif3']
                ws.cell(row=cont, column=45).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=46).value = paqNinio['fields']['dif4']
                ws.cell(row=cont, column=46).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=47).value = paqNinio['fields']['dif5']
                ws.cell(row=cont, column=47).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=48).value = paqNinio['fields']['dif6']
                ws.cell(row=cont, column=48).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=49).value = paqNinio['fields']['dif7']
                ws.cell(row=cont, column=49).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=50).value = paqNinio['fields']['dif8']
                ws.cell(row=cont, column=50).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=51).value = paqNinio['fields']['dif9']
                ws.cell(row=cont, column=51).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=52).value = paqNinio['fields']['dif10']
                ws.cell(row=cont, column=52).alignment = Alignment(horizontal="center")

                if paqNinio['fields']['num'] == '1':
                    cumplen = '✔'
                    ws.cell(row=cont, column=53).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=53).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=53).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=53).value = cumplen

                cont = cont+1
                num = num+1

        # sheet2 = wb.create_sheet('RESUMEN')
        # sheet2['A1'] = 'SUSCRIPCION'
        nombre_archivo = "DEIT_PASCO AVANCE PAQUETE NIÑO COMPLETO.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL PAQUETE NIÑO'
        wb.save(response)
        return response


class FollowPregnantView(TemplateView):
    template_name = 'pregnant/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.session['sytem']['typeca'] == 'CA':
            context['establecimiento'] = Establecimiento.objects.filter(codigo=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DS':
            context['establecimiento'] = Establecimiento.objects.filter(dist_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'PR':
            context['establecimiento'] = Establecimiento.objects.filter(prov_id=self.request.session['sytem']['codeca'])
        elif self.request.session['sytem']['typeca'] == 'DP':
            context['establecimiento'] = Establecimiento.objects.filter(dep_id=self.request.session['sytem']['codeca'])
        return context


class ListPregnantFollow(View):
    http_method_names = ['get', 'post', 'put', 'delete']
    def post(self,request,*args, **kwargs):
        json_data4 = []
        if len(request.POST['mes']) == 1:
            mes = '0'+request.POST['mes']
        else:
            mes = request.POST['mes']

        if request.POST['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                total = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                if total==0 or total==None:
                    total=0
                if cumplen==0 or cumplen==None:
                    cumplen=0

                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif self.request.session['sytem']['typeca'] == 'DS':
                print('AQUI TOY')
                print(self.request.session['sytem']['codeca'])
                total = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                if total==0 or total==None:
                    total=0
                if cumplen==0 or cumplen==None:
                    cumplen=0

                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dist=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif self.request.session['sytem']['typeca'] == 'PR':
                total = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                if total==0 or total==None:
                    total=0
                if cumplen==0 or cumplen==None:
                    cumplen=0

                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_prov=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif self.request.session['sytem']['typeca'] == 'DP':
                total = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).aggregate(total=Sum('den'))['total']
                cumplen = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).aggregate(cumplen=Sum('num'))['cumplen']
                if total==0 or total==None:
                    total=0
                if cumplen==0 or cumplen==None:
                    cumplen=0

                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
                dataProv = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).values('establecimiento').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_dep=self.request.session['sytem']['codeca']).order_by('cod_eess')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            total = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).aggregate(total=Count('id'))['total']
            cumplen = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', num=1, cod_eess=request.POST['eess']).aggregate(cumplen=Count('id'))['cumplen']
            if total==0 or total==None:
                total=0
            if cumplen==0 or cumplen==None:
                cumplen=0

            dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }
            dataProv = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).values('provincia').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataDist = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).values('establecimiento').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataNom = PregnantFollow.objects.filter(ctrl1__gte=request.POST['anio']+'-'+mes+'-01', cod_eess=request.POST['eess']).order_by('cod_eess')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)
        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintPackPregnant(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:AB2", "medium", "57267C")
        set_border(self, ws, "A4:AB4", "medium", "366092")
        set_border(self, ws, "A6:AB6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws.merge_cells('B2:BY2')
        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 6
        ws.column_dimensions['J'].width = 6
        ws.column_dimensions['K'].width = 11
        ws.column_dimensions['L'].width = 6
        ws.column_dimensions['M'].width = 11
        ws.column_dimensions['N'].width = 11
        ws.column_dimensions['O'].width = 11
        ws.column_dimensions['P'].width = 11
        ws.column_dimensions['Q'].width = 11
        ws.column_dimensions['R'].width = 11
        ws.column_dimensions['S'].width = 11
        ws.column_dimensions['T'].width = 11
        ws.column_dimensions['U'].width = 11
        ws.column_dimensions['V'].width = 11
        ws.column_dimensions['W'].width = 11
        ws.column_dimensions['X'].width = 11
        ws.column_dimensions['Y'].width = 11
        ws.column_dimensions['Z'].width = 11
        ws.column_dimensions['AA'].width = 11
        ws.column_dimensions['AB'].width = 11
        ws.column_dimensions['AC'].width = 11
        ws.column_dimensions['AD'].width = 11
        ws.column_dimensions['AE'].width = 11
        ws.column_dimensions['AF'].width = 11
        ws.column_dimensions['AG'].width = 11
        ws.column_dimensions['AH'].width = 11
        ws.column_dimensions['AI'].width = 11
        ws.column_dimensions['AJ'].width = 11
        ws.column_dimensions['AK'].width = 6
        ws.column_dimensions['AL'].width = 6
        ws.column_dimensions['AM'].width = 6
        ws.column_dimensions['AN'].width = 6
        ws.column_dimensions['AO'].width = 6

        ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'ESSALUD: Seguimiento de gestantes que recibieron el paquete integrado de servicios - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('A4:BY4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['A4'] = 'CODIFICACION: '

        ws.merge_cells('A6:BY6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: EsSalud con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A9'] = '#'
        ws['A9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['B9'] = 'Centro Asistencial'
        ws['B9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['C9'] = 'Documento'
        ws['C9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['D9'] = 'Apellidos y Nombres'
        ws['D9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['E9'] = 'Edad Capt.'
        ws['E9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['F9'] = 'Max Sem 13'
        ws['F9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['F9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['G9'] = 'Sem Captación'
        ws['G9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['G9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['H9'] = 'FUR'
        ws['H9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['H9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['I9'] = 'Peso'
        ws['I9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['J9'] = 'Talla'
        ws['J9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['K9'] = 'FPP'
        ws['K9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['L9'] = 'LBT'
        ws['L9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['M9'] = 'Result.'
        ws['M9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['N9'] = 'Tmz Viol.'
        ws['N9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N9'].fill = PatternFill(start_color='c7d2f0', end_color='c7d2f0', fill_type='solid')

        ws['O9'] = 'Ctrl 1'
        ws['O9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws.merge_cells('P8:Q8')
        ws['P8'] = 'Anemia'
        ws['P8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P8'].fill = PatternFill(start_color='bbe5ea', end_color='bbe5ea', fill_type='solid')

        ws['P9'] = 'Dx Anemia'
        ws['P9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P9'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws['Q9'] = 'Dx Anemia2'
        ws['Q9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Q9'].fill = PatternFill(start_color='c7ecf0', end_color='c7ecf0', fill_type='solid')

        ws.merge_cells('R8:S8')
        ws['R8'] = 'Riesgo Obstétrico'
        ws['R8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R8'].fill = PatternFill(start_color='d0c0ec', end_color='d0c0ec', fill_type='solid')

        ws['R9'] = 'ARO'
        ws['R9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['R9'].fill = PatternFill(start_color='d6c7f0', end_color='d6c7f0', fill_type='solid')

        ws['S9'] = 'BRO'
        ws['S9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['S9'].fill = PatternFill(start_color='d6c7f0', end_color='d6c7f0', fill_type='solid')

        ws.merge_cells('T8:V8')
        ws['T8'] = 'Consulta Médica'
        ws['T8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T8'].fill = PatternFill(start_color='dbbaea', end_color='dbbaea', fill_type='solid')

        ws['T9'] = '1 CMP'
        ws['T9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T9'].fill = PatternFill(start_color='e3c7f0', end_color='e3c7f0', fill_type='solid')

        ws['U9'] = '2 CMP'
        ws['U9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U9'].fill = PatternFill(start_color='e3c7f0', end_color='e3c7f0', fill_type='solid')

        ws['V9'] = '3 CMP'
        ws['V9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['V9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V9'].fill = PatternFill(start_color='e3c7f0', end_color='e3c7f0', fill_type='solid')

        ws.merge_cells('W8:AB8')
        ws['W8'] = 'Nutrición'
        ws['W8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W8'].fill = PatternFill(start_color='ecbbea', end_color='ecbbea', fill_type='solid')

        ws['W9'] = 'Nutri. 1'
        ws['W9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['W9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['W9'].fill = PatternFill(start_color='f0c7ee', end_color='f0c7ee', fill_type='solid')

        ws['X9'] = 'Nutri. 2'
        ws['X9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['X9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['X9'].fill = PatternFill(start_color='f0c7ee', end_color='f0c7ee', fill_type='solid')

        ws['Y9'] = 'Nutri. 3'
        ws['Y9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Y9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Y9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Y9'].fill = PatternFill(start_color='f0c7ee', end_color='f0c7ee', fill_type='solid')

        ws['Z9'] = 'Nutri. 4'
        ws['Z9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Z9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Z9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Z9'].fill = PatternFill(start_color='f0c7ee', end_color='f0c7ee', fill_type='solid')

        ws['AA9'] = 'Nutri. 5'
        ws['AA9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AA9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AA9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AA9'].fill = PatternFill(start_color='f0c7ee', end_color='f0c7ee', fill_type='solid')

        ws['AB9'] = 'Nutri. 6'
        ws['AB9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AB9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AB9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AB9'].fill = PatternFill(start_color='f0c7ee', end_color='f0c7ee', fill_type='solid')

        ws.merge_cells('AC8:AO8')
        ws['AC8'] = 'Atención Pre Natal'
        ws['AC8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AC8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AC8'].fill = PatternFill(start_color='d0e3f3', end_color='d0e3f3', fill_type='solid')

        ws['AC9'] = 'Control2'
        ws['AC9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AC9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AC9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AC9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AD9'] = 'Control3'
        ws['AD9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AD9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AD9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AD9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AE9'] = 'Control4'
        ws['AE9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AE9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AE9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AE9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AF9'] = 'Control5'
        ws['AF9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AF9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AF9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AF9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AG9'] = 'Control6'
        ws['AG9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AG9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AG9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AG9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AH9'] = 'Control7'
        ws['AH9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AH9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AH9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AH9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AI9'] = 'Control8'
        ws['AI9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AI9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AI9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AI9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AJ9'] = 'Control9'
        ws['AJ9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AJ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AJ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AJ9'].fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')

        ws['AK9'] = 'C1-C2'
        ws['AK9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AK9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AK9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AK9'].fill = PatternFill(start_color='d0e3f3', end_color='d0e3f3', fill_type='solid')

        ws['AL9'] = 'C2-C3'
        ws['AL9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AL9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AL9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AL9'].fill = PatternFill(start_color='d0e3f3', end_color='d0e3f3', fill_type='solid')

        ws['AM9'] = 'C3-C4'
        ws['AM9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AM9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AM9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AM9'].fill = PatternFill(start_color='d0e3f3', end_color='d0e3f3', fill_type='solid')

        ws['AN9'] = 'C4-C5'
        ws['AN9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AN9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AN9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AN9'].fill = PatternFill(start_color='d0e3f3', end_color='d0e3f3', fill_type='solid')

        ws['AO9'] = 'C5-C6'
        ws['AO9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AO9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AO9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AO9'].fill = PatternFill(start_color='d0e3f3', end_color='d0e3f3', fill_type='solid')

        ws.merge_cells('AP8:AQ8')
        ws['AP8'] = 'Odontología'
        ws['AP8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AP8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AP8'].fill = PatternFill(start_color='ebecc7', end_color='ebecc7', fill_type='solid')

        ws['AP9'] = 'Odonto. 1'
        ws['AP9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AP9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AP9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AP9'].fill = PatternFill(start_color='f1f2d0', end_color='f1f2d0', fill_type='solid')

        ws['AQ9'] = 'Odonto. 2'
        ws['AQ9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AQ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AQ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AQ9'].fill = PatternFill(start_color='f1f2d0', end_color='f1f2d0', fill_type='solid')

        ws['AR9'] = 'Psicología'
        ws['AR9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AR9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AR9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AR9'].fill = PatternFill(start_color='e1e2e3', end_color='e1e2e3', fill_type='solid')

        ws.merge_cells('AS8:AV8')
        ws['AS8'] = 'Enfermería'
        ws['AS8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AS8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AS8'].fill = PatternFill(start_color='c2ece4', end_color='c2ece4', fill_type='solid')

        ws['AS9'] = 'DT'
        ws['AS9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AS9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AS9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AS9'].fill = PatternFill(start_color='d0f2eb', end_color='d0f2eb', fill_type='solid')

        ws['AT9'] = 'DTPA'
        ws['AT9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AT9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AT9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AT9'].fill = PatternFill(start_color='d0f2eb', end_color='d0f2eb', fill_type='solid')

        ws['AU9'] = 'Hepatitis'
        ws['AU9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AU9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AU9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AU9'].fill = PatternFill(start_color='d0f2eb', end_color='d0f2eb', fill_type='solid')

        ws['AV9'] = 'Influenza'
        ws['AV9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AV9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AV9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AV9'].fill = PatternFill(start_color='d0f2eb', end_color='d0f2eb', fill_type='solid')

        ws['AW9'] = 'Ate Parto'
        ws['AW9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AW9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AW9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AW9'].fill = PatternFill(start_color='d3d0f2', end_color='d3d0f2', fill_type='solid')

        ws.merge_cells('AX8:BE8')
        ws['AX8'] = 'Suplementación Sulfato Ferroso'
        ws['AX8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AX8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AX8'].fill = PatternFill(start_color='d2e3c7', end_color='d2e3c7', fill_type='solid')

        ws['AX9'] = 'Suple 1'
        ws['AX9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AX9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AX9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AX9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['AY9'] = 'Suple 2'
        ws['AY9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AY9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AY9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AY9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['AZ9'] = 'Suple 3'
        ws['AZ9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AZ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AZ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AZ9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['BA9'] = 'Suple 4'
        ws['BA9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BA9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BA9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BA9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['BB9'] = 'Suple 5'
        ws['BB9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BB9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BB9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BB9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['BC9'] = 'Suple Ant 13'
        ws['BC9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BC9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BC9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BC9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['BD9'] = 'Suple Ant 13_2'
        ws['BD9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BD9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BD9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BD9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws['BE9'] = 'Suple Ant 13_3'
        ws['BE9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BE9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BE9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BE9'].fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')

        ws.merge_cells('BF8:BJ8')
        ws['BF8'] = 'Suple. Carbonato de Calcio'
        ws['BF8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BF8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BF8'].fill = PatternFill(start_color='e2dfc6', end_color='e2dfc6', fill_type='solid')

        ws['BF9'] = 'CaCO 1'
        ws['BF9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BF9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BF9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BF9'].fill = PatternFill(start_color='f0eeda', end_color='f0eeda', fill_type='solid')

        ws['BG9'] = 'CaCO 2'
        ws['BG9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BG9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BG9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BG9'].fill = PatternFill(start_color='f0eeda', end_color='f0eeda', fill_type='solid')

        ws['BH9'] = 'CaCO 3'
        ws['BH9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BH9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BH9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BH9'].fill = PatternFill(start_color='f0eeda', end_color='f0eeda', fill_type='solid')

        ws['BI9'] = 'CaCO 4'
        ws['BI9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BI9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BI9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BI9'].fill = PatternFill(start_color='f0eeda', end_color='f0eeda', fill_type='solid')

        ws['BJ9'] = 'CaCO 5'
        ws['BJ9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BJ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BJ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BJ9'].fill = PatternFill(start_color='f0eeda', end_color='f0eeda', fill_type='solid')

        ws.merge_cells('BK8:BP8')
        ws['BK8'] = 'Psicoprofilaxis'
        ws['BK8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BK8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BK8'].fill = PatternFill(start_color='dddaa1', end_color='dddaa1', fill_type='solid')

        ws['BK9'] = 'PPO1'
        ws['BK9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BK9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BK9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BK9'].fill = PatternFill(start_color='e8e5b0', end_color='e8e5b0', fill_type='solid')

        ws['BL9'] = 'PPO2'
        ws['BL9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BL9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BL9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BL9'].fill = PatternFill(start_color='e8e5b0', end_color='e8e5b0', fill_type='solid')

        ws['BM9'] = 'PPO3'
        ws['BM9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BM9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BM9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BM9'].fill = PatternFill(start_color='e8e5b0', end_color='e8e5b0', fill_type='solid')

        ws['BN9'] = 'PPO4'
        ws['BN9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BN9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BN9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BN9'].fill = PatternFill(start_color='e8e5b0', end_color='e8e5b0', fill_type='solid')

        ws['BO9'] = 'PPO5'
        ws['BO9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BO9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BO9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BO9'].fill = PatternFill(start_color='e8e5b0', end_color='e8e5b0', fill_type='solid')

        ws['BP9'] = 'PPO6'
        ws['BP9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BP9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BP9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BP9'].fill = PatternFill(start_color='e8e5b0', end_color='e8e5b0', fill_type='solid')

        ws.merge_cells('BQ8:BV8')
        ws['BQ8'] = 'Estimulación Prenatal'
        ws['BQ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BQ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BQ8'].fill = PatternFill(start_color='efedbb', end_color='efedbb', fill_type='solid')

        ws['BQ9'] = 'EPN1'
        ws['BQ9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BQ9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BQ9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BQ9'].fill = PatternFill(start_color='f5f3cb', end_color='f5f3cb', fill_type='solid')

        ws['BR9'] = 'EPN2'
        ws['BR9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BR9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BR9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BR9'].fill = PatternFill(start_color='f5f3cb', end_color='f5f3cb', fill_type='solid')

        ws['BS9'] = 'EPN3'
        ws['BS9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BS9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BS9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BS9'].fill = PatternFill(start_color='f5f3cb', end_color='f5f3cb', fill_type='solid')

        ws['BT9'] = 'EPN4'
        ws['BT9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BT9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BT9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BT9'].fill = PatternFill(start_color='f5f3cb', end_color='f5f3cb', fill_type='solid')

        ws['BU9'] = 'EPN5'
        ws['BU9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BU9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BU9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BU9'].fill = PatternFill(start_color='f5f3cb', end_color='f5f3cb', fill_type='solid')

        ws['BV9'] = 'EPN6'
        ws['BV9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BV9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BV9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BV9'].fill = PatternFill(start_color='f5f3cb', end_color='f5f3cb', fill_type='solid')

        ws['BW9'] = 'Ini Sem 28'
        ws['BW9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BW9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BW9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BW9'].fill = PatternFill(start_color='ebf5b3', end_color='ebf5b3', fill_type='solid')

        ws['BX9'] = 'Ini Sem 33'
        ws['BX9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BX9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BX9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BX9'].fill = PatternFill(start_color='ebf5b3', end_color='ebf5b3', fill_type='solid')

        ws['BY9'] = 'Cumple'
        ws['BY9'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BY9'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BY9'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BY9'].fill = PatternFill(start_color='b3f5c2', end_color='b3f5c2', fill_type='solid')

        if len(request.GET['mes']) == 1:
            mes = '0'+request.GET['mes']
        else:
            mes = request.GET['mes']

        if request.GET['eess'] == 'TODOS':
            if self.request.session['sytem']['typeca'] == 'CA':
                dataNom = PregnantFollow.objects.filter(cod_eess=self.request.session['sytem']['codeca'], ctrl1__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'DS':
                dataNom = PregnantFollow.objects.filter(cod_dist=self.request.session['sytem']['codeca'], ctrl1__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'PR':
                dataNom = PregnantFollow.objects.filter(cod_prov=self.request.session['sytem']['codeca'], ctrl1__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            elif self.request.session['sytem']['typeca'] == 'DP':
                dataNom = PregnantFollow.objects.filter(cod_dep=self.request.session['sytem']['codeca'], ctrl1__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
        else:
            dataNom = PregnantFollow.objects.filter(cod_eess=request.GET['eess'], ctrl1__gte=request.GET['anio']+'-'+mes+'-01').order_by('cod_eess')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 10
        cant = len(dataNom)
        num=1
        if cant > 0:
            for paqGest in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = paqGest['fields']['establecimiento']
                ws.cell(row=cont, column=3).value = paqGest['fields']['documento']
                ws.cell(row=cont, column=4).value = paqGest['fields']['ape_nombres']
                ws.cell(row=cont, column=5).value = paqGest['fields']['edad_cap']
                ws.cell(row=cont, column=6).value = paqGest['fields']['max_sem13']
                ws.cell(row=cont, column=7).value = paqGest['fields']['sem_captada']
                ws.cell(row=cont, column=8).value = paqGest['fields']['fur']
                ws.cell(row=cont, column=9).value = paqGest['fields']['peso']
                ws.cell(row=cont, column=10).value = paqGest['fields']['talla']
                ws.cell(row=cont, column=11).value = paqGest['fields']['fpp']
                ws.cell(row=cont, column=12).value = paqGest['fields']['laboratorio']
                ws.cell(row=cont, column=13).value = paqGest['fields']['result']
                ws.cell(row=cont, column=14).value = paqGest['fields']['tmz']
                ws.cell(row=cont, column=15).value = paqGest['fields']['ctrl1']
                ws.cell(row=cont, column=16).value = paqGest['fields']['dx_anemia']
                ws.cell(row=cont, column=17).value = paqGest['fields']['dx_anemia2']
                ws.cell(row=cont, column=18).value = paqGest['fields']['aro']
                ws.cell(row=cont, column=19).value = paqGest['fields']['bro']
                ws.cell(row=cont, column=20).value = paqGest['fields']['cmp1']
                ws.cell(row=cont, column=21).value = paqGest['fields']['cmp2']
                ws.cell(row=cont, column=22).value = paqGest['fields']['cmp3']
                ws.cell(row=cont, column=23).value = paqGest['fields']['nutricion1']
                ws.cell(row=cont, column=24).value = paqGest['fields']['nutricion2']
                ws.cell(row=cont, column=25).value = paqGest['fields']['nutricion3']
                ws.cell(row=cont, column=26).value = paqGest['fields']['nutricion4']
                ws.cell(row=cont, column=27).value = paqGest['fields']['nutricion5']
                ws.cell(row=cont, column=28).value = paqGest['fields']['nutricion6']
                ws.cell(row=cont, column=29).value = paqGest['fields']['ctrl2']
                ws.cell(row=cont, column=30).value = paqGest['fields']['ctrl3']
                ws.cell(row=cont, column=31).value = paqGest['fields']['ctrl4']
                ws.cell(row=cont, column=32).value = paqGest['fields']['ctrl5']
                ws.cell(row=cont, column=33).value = paqGest['fields']['ctrl6']
                ws.cell(row=cont, column=34).value = paqGest['fields']['ctrl7']
                ws.cell(row=cont, column=35).value = paqGest['fields']['ctrl8']
                ws.cell(row=cont, column=36).value = paqGest['fields']['ctrl9']
                ws.cell(row=cont, column=37).value = paqGest['fields']['c1_c2']
                ws.cell(row=cont, column=38).value = paqGest['fields']['c2_c3']
                ws.cell(row=cont, column=39).value = paqGest['fields']['c3_c4']
                ws.cell(row=cont, column=40).value = paqGest['fields']['c4_c5']
                ws.cell(row=cont, column=41).value = paqGest['fields']['c5_c6']

                ws.cell(row=cont, column=42).value = paqGest['fields']['odontologia1']
                ws.cell(row=cont, column=43).value = paqGest['fields']['odontologia2']
                ws.cell(row=cont, column=44).value = paqGest['fields']['psicologia']
                ws.cell(row=cont, column=45).value = paqGest['fields']['dt']
                ws.cell(row=cont, column=46).value = paqGest['fields']['dtpa']
                ws.cell(row=cont, column=47).value = paqGest['fields']['hepatitis']
                ws.cell(row=cont, column=48).value = paqGest['fields']['influenza']
                ws.cell(row=cont, column=49).value = paqGest['fields']['ate_parto']
                ws.cell(row=cont, column=50).value = paqGest['fields']['suple1']
                ws.cell(row=cont, column=51).value = paqGest['fields']['suple2']
                ws.cell(row=cont, column=52).value = paqGest['fields']['suple3']
                ws.cell(row=cont, column=53).value = paqGest['fields']['suple4']
                ws.cell(row=cont, column=54).value = paqGest['fields']['suple5']
                ws.cell(row=cont, column=55).value = paqGest['fields']['suple_ant13']
                ws.cell(row=cont, column=56).value = paqGest['fields']['suple_ant13_2']
                ws.cell(row=cont, column=57).value = paqGest['fields']['suple_ant13_3']
                ws.cell(row=cont, column=58).value = paqGest['fields']['carbonato1']
                ws.cell(row=cont, column=59).value = paqGest['fields']['carbonato2']
                ws.cell(row=cont, column=60).value = paqGest['fields']['carbonato3']
                ws.cell(row=cont, column=61).value = paqGest['fields']['carbonato4']
                ws.cell(row=cont, column=62).value = paqGest['fields']['carbonato5']
                ws.cell(row=cont, column=63).value = paqGest['fields']['ppo1']
                ws.cell(row=cont, column=64).value = paqGest['fields']['ppo2']
                ws.cell(row=cont, column=65).value = paqGest['fields']['ppo3']
                ws.cell(row=cont, column=66).value = paqGest['fields']['ppo4']
                ws.cell(row=cont, column=67).value = paqGest['fields']['ppo5']
                ws.cell(row=cont, column=68).value = paqGest['fields']['ppo6']
                ws.cell(row=cont, column=69).value = paqGest['fields']['epn1']
                ws.cell(row=cont, column=70).value = paqGest['fields']['epn2']
                ws.cell(row=cont, column=71).value = paqGest['fields']['epn3']
                ws.cell(row=cont, column=72).value = paqGest['fields']['epn4']
                ws.cell(row=cont, column=73).value = paqGest['fields']['epn5']
                ws.cell(row=cont, column=74).value = paqGest['fields']['epn6']
                ws.cell(row=cont, column=75).value = paqGest['fields']['ini_sem28']
                ws.cell(row=cont, column=76).value = paqGest['fields']['ini_sem33']

                if paqGest['fields']['num'] == '1':
                    cumplen = '✔'
                    ws.cell(row=cont, column=77).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=77).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=77).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=77).value = cumplen

                cont = cont+1
                num = num+1

        # sheet2 = wb.create_sheet('RESUMEN')
        # sheet2['A1'] = 'SUSCRIPCION'
        nombre_archivo = "DEIT_PASCO AVANCE PAQUETE GESTANTE.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL PAQUETE GESTANTE'
        wb.save(response)
        return response
